import sys
from pathlib import Path
from datetime import datetime, date, timedelta

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QMessageBox, QFrame
)

# openpyxl은 .xls 저장/편집 불가 -> 안전하게 .xlsx만
ALLOWED_EXT = {".xlsx"}


# =========================
# 공통 유틸
# =========================
def norm(v) -> str:
    return str(v).replace("\n", " ").strip().lower()


def find_col_by_keywords_ws(ws, header_row: int, keywords: list[str], mode: str = "any") -> int:
    """
    ws에서 header_row를 기준으로 keywords로 컬럼 찾기(1-index)

    mode:
      - "any": keywords 중 하나라도 포함되면 매칭(OR)
      - "all": keywords 전부 포함되어야 매칭(AND)
    """
    mode = (mode or "any").lower()
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v in (None, ""):
            continue
        s = norm(v)

        if mode == "all":
            ok = all(k.lower() in s for k in keywords)
        else:
            ok = any(k.lower() in s for k in keywords)

        if ok:
            return col

    raise ValueError(f"컬럼을 찾을 수 없습니다: {keywords} (mode={mode})")


def parse_int_like(v):
    """주행거리 등 숫자 파싱(콤마/문자 섞여도 최대한)"""
    if v in (None, ""):
        return None
    try:
        if isinstance(v, (int, float)):
            return int(float(v))
        s = str(v).strip().replace(",", "")
        if not s:
            return None
        return int(float(s))
    except:
        return None


def parse_excel_date(v):
    """
    파일별 날짜 표기가 달라도 최대한 파싱:
    - datetime/date
    - 20250725 같은 int/str(yyyymmdd)
    - 2025-07-25, 2025/07/25, 2025.07.25 등
    - 엑셀 시리얼(대략 20000~80000 범위)도 처리
    """
    if v in (None, ""):
        return None

    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    # numeric: yyyymmdd or excel serial
    if isinstance(v, (int, float)):
        iv = int(v)

        # yyyymmdd 추정
        if 19000101 <= iv <= 21001231:
            s = str(iv)
            try:
                return datetime.strptime(s, "%Y%m%d").date()
            except:
                pass

        # excel serial 추정(대략)
        if 20000 <= iv <= 80000:
            base = datetime(1899, 12, 30)  # Excel 관행
            try:
                return (base + timedelta(days=float(v))).date()
            except:
                return None

    # string parse
    s = str(v).strip()
    if not s:
        return None

    # yyyymmdd 문자열(구분자 제거)
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d").date()
        except:
            pass

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y %m %d",
                "%y-%m-%d", "%y/%m/%d", "%y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            continue

    return None


def guess_last_data_row(ws, data_start_row: int, anchor_col: int, empty_run: int = 30) -> int:
    """
    데이터 끝 추정:
    anchor_col(예: repair_date)을 기준으로 연속 empty가 empty_run 이상이면 그 전을 데이터 끝으로 봄
    """
    last = ws.max_row
    streak = 0
    for r in range(data_start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=anchor_col).value
        if v in (None, ""):
            streak += 1
            if streak >= empty_run:
                last = r - empty_run
                break
        else:
            streak = 0
    return max(last, data_start_row)


# =========================
# 1) 차계 병합 해제 + 채우기 (last_row까지만)
# =========================
def unmerge_and_fill_column(ws, target_col: int, data_start_row: int, last_row: int):
    merged_ranges = list(ws.merged_cells.ranges)

    # 병합 해제 + 병합범위 전체 채우기
    for mr in merged_ranges:
        if (mr.min_col <= target_col <= mr.max_col) and (mr.min_row >= data_start_row):
            top_left = ws.cell(mr.min_row, mr.min_col).value
            ws.unmerge_cells(str(mr))
            for r in range(mr.min_row, min(mr.max_row, last_row) + 1):
                ws.cell(row=r, column=target_col).value = top_left

    # 병합이 아닌 빈칸 ffill (last_row까지만)
    prev = None
    for r in range(data_start_row, last_row + 1):
        c = ws.cell(row=r, column=target_col)
        if c.value in (None, ""):
            if prev not in (None, ""):
                c.value = prev
        else:
            prev = c.value


# =========================
# 2) 구상율 변경(단일 진입점) + 바뀐 행 추적
# =========================
def set_rate(ws, row: int, rate_col: int, new_rate: float, changed_rows: set[int]):
    cell = ws.cell(row=row, column=rate_col)
    old = cell.value

    try:
        old_f = float(str(old).replace(",", "")) if old not in (None, "") else None
    except:
        old_f = None

    if old_f != float(new_rate):
        cell.value = float(new_rate)
        changed_rows.add(row)


# =========================
# 3) 바뀐 행만 구상금액 수식
# =========================
def set_chargeback_formula_rows(ws, rows: set[int], occ_col: int, rate_col: int, chb_col: int):
    for r in rows:
        occ_addr = ws.cell(row=r, column=occ_col).coordinate
        rate_addr = ws.cell(row=r, column=rate_col).coordinate
        ws.cell(row=r, column=chb_col).value = f"={occ_addr}*({rate_addr}/100)"


# =========================
# 4) 발생/구상 합계 행 추가
# =========================
def add_sum_rows(ws, data_start_row: int, last_row: int, occ_col: int, chb_col: int):
    sum_start_row = last_row + 3

    ws.cell(row=sum_start_row, column=occ_col - 1).value = "발생금액"
    ws.cell(row=sum_start_row, column=occ_col).value = (
        f"=SUM({ws.cell(row=data_start_row, column=occ_col).coordinate}:"
        f"{ws.cell(row=last_row, column=occ_col).coordinate})"
    )

    ws.cell(row=sum_start_row + 1, column=chb_col - 1).value = "구상금액"
    ws.cell(row=sum_start_row + 1, column=chb_col).value = (
        f"=SUM({ws.cell(row=data_start_row, column=chb_col).coordinate}:"
        f"{ws.cell(row=last_row, column=chb_col).coordinate})"
    )


# =========================
# 5) SUBTOTAL 수식
# =========================
def set_subtotal(ws, target_col: int, data_start_row: int, last_row: int, subtotal_row: int, use_109: bool = True):
    func = 109 if use_109 else 9
    ws.cell(row=subtotal_row, column=target_col).value = (
        f"=SUBTOTAL({func},"
        f"{ws.cell(row=data_start_row, column=target_col).coordinate}:"
        f"{ws.cell(row=last_row, column=target_col).coordinate})"
    )


# =========================
# 6) 마일리지/보증기간 필터 (색칠 + 구상율 0)
# =========================
def apply_warranty_filters_ws(
    ws,
    header_row: int,
    data_start_row: int,
    last_row: int,
    changed_rows: set[int],
    mileage_threshold: int = 50000,
    warranty_years: int = 2,
):
    fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    mileage_col = find_col_by_keywords_ws(ws, header_row, ["mileage", "주행거리"], mode="any")
    sale_col    = find_col_by_keywords_ws(ws, header_row, ["sale date", "판매일", "sale"], mode="any")
    repair_col  = find_col_by_keywords_ws(ws, header_row, ["repair date", "수리일자", "repair"], mode="any")
    rate_col    = find_col_by_keywords_ws(ws, header_row, ["구상율", "liability ratio", "ratio"], mode="any")

    warranty_days = int(warranty_years * 365)

    for r in range(data_start_row, last_row + 1):
        mv = parse_int_like(ws.cell(row=r, column=mileage_col).value)
        if mv is not None and mv >= mileage_threshold:
            ws.cell(row=r, column=mileage_col).fill = fill_color
            set_rate(ws, r, rate_col, 0, changed_rows)

        sale_dt = parse_excel_date(ws.cell(row=r, column=sale_col).value)
        repair_dt = parse_excel_date(ws.cell(row=r, column=repair_col).value)

        if sale_dt and repair_dt:
            if (repair_dt - sale_dt).days >= warranty_days:
                ws.cell(row=r, column=sale_col).fill = fill_color
                set_rate(ws, r, rate_col, 0, changed_rows)

    return rate_col


# =========================
# 전체 처리 파이프
# =========================
def process_all(in_path: str, out_path: str):
    sheet_index = 0
    header_row = 3
    data_start_row = 4

    wb = load_workbook(in_path)
    ws = wb.worksheets[sheet_index]

    vehicle_col = find_col_by_keywords_ws(ws, header_row, ["vehicle", "차계"], mode="any")
    occ_col     = find_col_by_keywords_ws(ws, header_row, ["total cost", "발생", "발생금액"], mode="any")
    chb_col     = find_col_by_keywords_ws(ws, header_row, ["chargeback", "구상", "구상금액"], mode="any")
    repair_col  = find_col_by_keywords_ws(ws, header_row, ["repair date", "수리일자", "repair"], mode="any")

    last_row = guess_last_data_row(ws, data_start_row, anchor_col=repair_col, empty_run=30)

    unmerge_and_fill_column(ws, vehicle_col, data_start_row, last_row)

    changed_rows: set[int] = set()

    rate_col = apply_warranty_filters_ws(
        ws=ws,
        header_row=header_row,
        data_start_row=data_start_row,
        last_row=last_row,
        changed_rows=changed_rows,
        mileage_threshold=50000,
        warranty_years=2,
    )

    set_chargeback_formula_rows(ws, changed_rows, occ_col, rate_col, chb_col)

    add_sum_rows(ws, data_start_row, last_row, occ_col, chb_col)

    # ✅ "겉보기 빈칸"까지 포함해서 SUBTOTAL 안전 삽입
    subtotal_row = header_row - 1
    v = ws.cell(row=subtotal_row, column=chb_col).value
    if v is None or str(v).strip() == "":
        set_subtotal(ws, chb_col, data_start_row, last_row, subtotal_row, use_109=True)

    wb.save(out_path)


# =========================
# GUI
# =========================
class DropZone(QFrame):
    def __init__(self, on_file_dropped):
        super().__init__()
        self.on_file_dropped = on_file_dropped
        self.setAcceptDrops(True)
        self.setFixedHeight(80)
        self.setStyleSheet("QFrame { border: 1px dashed #888; border-radius: 6px; }")
        lay = QVBoxLayout()
        self.lbl = QLabel("엑셀 드래그 (.xlsx)")
        self.lbl.setAlignment(Qt.AlignCenter)
        lay.addWidget(self.lbl)
        self.setLayout(lay)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            p = Path(event.mimeData().urls()[0].toLocalFile())
            if p.suffix.lower() in ALLOWED_EXT:
                event.acceptProposedAction()
                return
        event.ignore()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls:
            self.on_file_dropped(Path(urls[0].toLocalFile()))


class App(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("AMS")
        self.setFixedSize(420, 200)

        self.in_path: Path | None = None

        root = QVBoxLayout()

        self.drop = DropZone(self.set_file)
        root.addWidget(self.drop)

        row = QHBoxLayout()
        self.lbl = QLabel("파일: (없음)")
        self.lbl.setWordWrap(True)
        btn = QPushButton("선택")
        btn.clicked.connect(self.pick_file)
        row.addWidget(self.lbl, 1)
        row.addWidget(btn)
        root.addLayout(row)

        self.btn_export = QPushButton("저장")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self.export_processed)
        root.addWidget(self.btn_export)

        self.setLayout(root)

    def set_file(self, p: Path):
        if p.suffix.lower() not in ALLOWED_EXT:
            QMessageBox.warning(self, "확장자", "현재는 .xlsx만 지원합니다.")
            return
        self.in_path = p
        self.lbl.setText(f"파일: {p.name}")
        self.btn_export.setEnabled(True)

    def pick_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "엑셀 선택", "", "Excel Files (*.xlsx)")
        if path:
            self.set_file(Path(path))

    def export_processed(self):
        if not self.in_path:
            return

        ts = datetime.now().strftime("%H%M%S")
        default = f"{self.in_path.stem}_{ts}{self.in_path.suffix}"

        save_path, _ = QFileDialog.getSaveFileName(self, "저장", default, "Excel Files (*.xlsx)")
        if not save_path:
            return

        try:
            process_all(str(self.in_path), save_path)
            QMessageBox.information(self, "완료", "처리 후 저장됨")
        except Exception as e:
            QMessageBox.critical(self, "오류", str(e))


def main():
    app = QApplication(sys.argv)
    w = App()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
