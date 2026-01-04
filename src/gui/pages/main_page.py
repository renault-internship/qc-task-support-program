from __future__ import annotations

from pathlib import Path
from typing import Dict, Any

from PySide6.QtCore import Qt, QSortFilterProxyModel, QRegularExpression, QStringListModel, QModelIndex, QThread, Signal
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QFileDialog, QMessageBox,
    QAbstractItemView, QMenu, QSplitter, QDialog, QApplication
)

from openpyxl.workbook.workbook import Workbook

from src.utils import load_workbook_safe, save_workbook_safe, AppError
from src.excel_processor import preprocess_inplace
from src.database import (
    get_company_info, get_all_companies, get_all_companies_with_code,
    get_rules_from_table, add_rule_to_table
)
from src.gui.containers import (
    PreviewContainer, InfoPanel, ControlPanel
)
from src.gui.models import ExcelSheetModel
from src.gui.excel_filter import ExcelFilterProxyModel, ColumnFilterDialog, ColumnSelectDialog
from src.gui.dialogs import AddRuleDialog


class WorkerThread(QThread):
    """긴 작업을 처리할 백그라운드 쓰레드"""
    finished = Signal(object)
    error = Signal(str)

    def __init__(self, task_fn, *args, **kwargs):
        super().__init__()
        self.task_fn = task_fn
        self.args = args
        self.kwargs = kwargs

    def run(self):
        try:
            result = self.task_fn(*self.args, **self.kwargs)
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class MainPageWidget(QWidget):
    """메인 페이지 - 엑셀 전처리 도구"""

    def __init__(self, parent=None):
        super().__init__(parent)

        self.file_path_domestic: Path | None = None
        self.file_path_overseas: Path | None = None
        self.wb_domestic: Workbook | None = None
        self.wb_overseas: Workbook | None = None
        self.model: ExcelSheetModel | None = None
        self.proxy: QSortFilterProxyModel | None = None
        self.current_company_info: Dict[str, Any] | None = None
        # 전처리 상태 추적
        self.preprocessed_domestic: bool = False
        self.preprocessed_overseas: bool = False

        # ================= 컨테이너 생성 =================
        self.control_panel = ControlPanel(self)
        self.preview_container = PreviewContainer(self)
        self.info_panel = InfoPanel(self)

        # ================= 레이아웃 구성 =================
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.control_panel)
        layout.addSpacing(4)

        # QSplitter로 preview와 info_panel 높이 조절 가능하게
        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(self.preview_container)
        splitter.addWidget(self.info_panel)

        # 초기 stretch 비율
        splitter.setStretchFactor(0, 4)  # preview_container가 크게
        splitter.setStretchFactor(1, 1)  # info_panel 작게

        # handle 시각적으로 표시
        splitter.setHandleWidth(6)
        splitter.setStyleSheet("""
        QSplitter::handle {
            background-color: #DDD;
            border-top: 2px solid #AAA;
            border-bottom: 2px solid #AAA;
        }
        QSplitter::handle:hover {
            background-color: #BBB;
        }
        """)

        layout.addWidget(splitter, 1)

        self.setLayout(layout)

        self._connect_signals()
        self._initialize()

    # ================= 초기화 =================
    def _initialize(self):
        self.info_panel.set_remark("-")
        self.load_companies()
        # 초기 전처리 버튼 상태 설정
        self._update_preprocess_button_state()

    def _connect_signals(self):
        self.control_panel.get_upload_domestic_button().clicked.connect(lambda: self.open_file("domestic"))
        self.control_panel.get_upload_overseas_button().clicked.connect(lambda: self.open_file("overseas"))
        self.control_panel.get_preprocess_button().clicked.connect(self.on_preprocess_clicked)
        # 검색창에서 Enter 키 또는 편집 완료 시 회사 선택
        self.control_panel.get_company_edit().editingFinished.connect(self._on_company_search_finished)
        self.control_panel.get_company_edit().returnPressed.connect(self._on_company_search_finished)
        # 자동완성 목록에서 항목 선택 시
        self.control_panel.get_company_completer().activated.connect(self._on_company_selected_from_completer)
        self.control_panel.get_search_edit().textChanged.connect(self.on_search_changed)
        self.control_panel.get_edit_all_checkbox().stateChanged.connect(self.on_edit_mode_changed)
        
        # 실행취소/다시실행 버튼 연결
        self.control_panel.get_undo_button().clicked.connect(self.on_undo)
        self.control_panel.get_redo_button().clicked.connect(self.on_redo)

        self.control_panel.get_sheet_combo().currentTextChanged.connect(self.on_sheet_changed)
        self.control_panel.get_export_final_button().clicked.connect(self.save_as_file)
        self.control_panel.get_filter_button().clicked.connect(self.on_filter_button_clicked)
        self.control_panel.get_clear_filter_button().clicked.connect(self.on_clear_filter_clicked)

    # ================= 회사 =================
    def load_companies(self):
        """협력사 목록 로드 및 자동완성 설정 (코드와 이름 모두 포함)"""
        companies_data = get_all_companies_with_code()
        if companies_data:
            # 코드와 이름을 모두 포함한 리스트 생성
            # 형식: "이름 (코드)"만 사용
            company_list = []
            for company in companies_data:
                sap_code = company["sap_code"]
                sap_name = company["sap_name"]
                # "이름 (코드)" 형식만 추가
                company_list.append(f"{sap_name} ({sap_code})")
            
            # QCompleter에 모델 설정
            model = QStringListModel(company_list, self)
            completer = self.control_panel.get_company_completer()
            completer.setModel(model)

    def _on_company_search_finished(self):
        """검색창에서 Enter 키 또는 편집 완료 시 호출"""
        text = self.control_panel.get_company_edit().text().strip()
        # "이름 (코드)" 또는 "코드 - 이름" 형식에서 실제 이름 또는 코드 추출
        name_or_code = self._extract_company_name_or_code(text)
        self._on_company_changed(name_or_code)
    
    def _on_company_selected_from_completer(self, text: str):
        """자동완성 목록에서 항목 선택 시 호출"""
        # "이름 (코드)" 또는 "코드 - 이름" 형식에서 실제 이름 또는 코드 추출
        name_or_code = self._extract_company_name_or_code(text)
        self._on_company_changed(name_or_code)
    
    def _extract_company_name_or_code(self, text: str) -> str:
        """자동완성 텍스트에서 실제 회사 이름 또는 코드 추출"""
        text = text.strip()
        # "이름 (코드)" 형식인 경우
        if " (" in text and text.endswith(")"):
            # 이름 부분만 반환
            return text.split(" (")[0]
        # 그 외의 경우는 그대로 반환 (직접 입력한 경우: 이름 또는 코드)
        return text
    
    def _on_company_changed(self, name: str):
        """회사 선택 시 정보 로드"""
        if not name:
            self.info_panel.set_company_info("", "")
            self.info_panel.set_rules([])
            self.current_company_info = None
            return

        company_info = get_company_info(name)
        if not company_info:
            # 검색 결과가 없으면 경고만 표시하고 초기화하지 않음
            QMessageBox.warning(self, "오류", f"기업정보를 찾을 수 없습니다: {name}")
            return

        self.current_company_info = company_info

        # 회사 정보
        self.info_panel.set_company_info(
            company_info.get("sap_name", ""),
            company_info.get("sap_code", "")
        )

        # Remark
        self.info_panel.set_remark(company_info.get("remark", ""))

        # Rule → InfoPanel에 바로 표시
        rule_table_name = company_info.get("rule_table_name")
        if rule_table_name:
            rules = get_rules_from_table(rule_table_name)
            self.info_panel.set_rules(rules)
        else:
            self.info_panel.set_rules([])

    # ================= Rule 추가 =================
    def add_rule(self):
        if not self.current_company_info:
            QMessageBox.warning(self, "오류", "먼저 기업을 선택해주세요.")
            return

        rule_table_name = self.current_company_info.get("rule_table_name")
        if not rule_table_name:
            QMessageBox.warning(self, "오류", "선택한 기업에 Rule 테이블이 없습니다.")
            return

        dialog = AddRuleDialog(rule_table_name, self)
        if dialog.exec():
            try:
                add_rule_to_table(rule_table_name=rule_table_name, **dialog.get_data())
                QMessageBox.information(self, "완료", "규칙이 추가되었습니다.")
                self._on_company_changed(
                    self.control_panel.get_company_edit().text().strip()
                )
            except Exception as e:
                QMessageBox.critical(self, "오류", str(e))

    # ================= 업로드 =================
    def open_file(self, file_type: str = "domestic"):
        """
        파일 업로드
        Args:
            file_type: "domestic" (국내) 또는 "overseas" (해외)
        """
        title = "국내 청구서 선택" if file_type == "domestic" else "해외 청구서 선택"
        path, _ = QFileDialog.getOpenFileName(
            self, title, "", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        file_path = Path(path)
        
        # 로딩 애니메이션 표시
        self.preview_container.show_loading("파일을 불러오는 중")
        
        # 백그라운드에서 파일 로드 실행
        self.load_worker = WorkerThread(load_workbook_safe, file_path)
        self.load_worker.finished.connect(lambda wb: self._on_load_finished(wb, file_type, file_path))
        self.load_worker.error.connect(self._on_worker_error)
        self.load_worker.start()

    def _on_load_finished(self, wb, file_type, file_path):
        """파일 로드 완료 시 호출되는 콜백"""
        # 워크북 저장
        if file_type == "domestic":
            self.file_path_domestic = file_path
            self.wb_domestic = wb
            self.preprocessed_domestic = False  # 새로 불러오면 전처리 상태 초기화
        else:
            self.file_path_overseas = file_path
            self.wb_overseas = wb
            self.preprocessed_overseas = False  # 새로 불러오면 전처리 상태 초기화

        # 시트 목록 업데이트
        self._update_sheet_list()
        QApplication.processEvents()

        # 불러온 파일 타입에 맞는 첫 번째 시트 로드
        sheet_combo = self.control_panel.get_sheet_combo()
        if sheet_combo.count() > 0:
            # 불러온 파일 타입에 맞는 시트 찾기
            target_prefix = "국내: " if file_type == "domestic" else "해외: "
            found_index = -1
            for i in range(sheet_combo.count()):
                if sheet_combo.itemText(i).startswith(target_prefix):
                    found_index = i
                    break
            
            if found_index >= 0:
                sheet_combo.setCurrentIndex(found_index)
                self._load_sheet_from_combo()
            else:
                # 찾지 못한 경우 첫 번째 시트 로드
                sheet_combo.setCurrentIndex(0)
                self._load_sheet_from_combo()

        # 전처리 버튼 상태 업데이트
        self._update_preprocess_button_state()
        
        remark = f"{'국내' if file_type == 'domestic' else '해외'} 청구서 업로드 완료. 전처리 전 상태"
        self.info_panel.set_remark(remark)
        
        # 모든 처리가 끝난 후 로딩 애니메이션 숨김
        QApplication.processEvents()
        self.preview_container.hide_loading()
    
    def _update_sheet_list(self):
        """시트 목록 업데이트 (국내/해외 모두 포함)"""
        sheet_combo = self.control_panel.get_sheet_combo()
        sheet_combo.blockSignals(True)
        sheet_combo.clear()
        
        # 국내 청구서 시트 추가
        if self.wb_domestic:
            for sheet_name in self.wb_domestic.sheetnames:
                sheet_combo.addItem(f"국내: {sheet_name}")
        
        # 해외 청구서 시트 추가
        if self.wb_overseas:
            for sheet_name in self.wb_overseas.sheetnames:
                sheet_combo.addItem(f"해외: {sheet_name}")
        
        sheet_combo.blockSignals(False)
    
    def _load_sheet_from_combo(self):
        """시트 콤보박스에서 선택한 시트 로드"""
        sheet_combo = self.control_panel.get_sheet_combo()
        sheet_text = sheet_combo.currentText()
        if not sheet_text:
            return
        self.load_sheet(sheet_text)

    # ================= 시트 =================
    def load_sheet(self, sheet_display_name: str):
        """
        시트 로드
        Args:
            sheet_display_name: "국내: Sheet1" 또는 "해외: Sheet1" 형식
        """
        # "국내: Sheet1" 또는 "해외: Sheet1" 형식에서 파싱
        if sheet_display_name.startswith("국내: "):
            if not self.wb_domestic:
                return
            actual_sheet_name = sheet_display_name.replace("국내: ", "")
            wb = self.wb_domestic
        elif sheet_display_name.startswith("해외: "):
            if not self.wb_overseas:
                return
            actual_sheet_name = sheet_display_name.replace("해외: ", "")
            wb = self.wb_overseas
        else:
            # 기존 형식 호환성 (없을 수도 있음)
            return

        if actual_sheet_name not in wb.sheetnames:
            return

        ws = wb[actual_sheet_name]
        self.model = ExcelSheetModel(ws, parent=self)

        edit_all = self.control_panel.get_edit_all_checkbox().isChecked()
        self.model.set_edit_all(edit_all)

        self.proxy = ExcelFilterProxyModel(self)
        self.proxy.setSourceModel(self.model)
        self.proxy.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.proxy.setFilterKeyColumn(-1)
        
        # model에 proxy 참조 설정 (SUBTOTAL 계산 시 필터 상태 확인용)
        self.model.set_proxy_model(self.proxy)

        table = self.preview_container.get_table()
        table.clearSpans()
        table.setModel(self.proxy)
        
        # setModel 후 delegate 다시 설정 (말줄임표 방지)
        from src.gui.containers.preview_container import NoElideDelegate
        table.setItemDelegate(NoElideDelegate(table))

        table.setAlternatingRowColors(True)
        table.setSortingEnabled(False)  # 컬럼 헤더 클릭 정렬 비활성화
        table.setSelectionBehavior(QAbstractItemView.SelectItems)
        table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        
        # 필터 상태 업데이트
        self._update_filter_button_state()
        
        # Undo/Redo 버튼 상태 업데이트
        self._update_undo_redo_buttons()
        
        # 모델의 dataChanged 시그널에 연결하여 편집 시 버튼 상태 업데이트
        if self.model:
            self.model.dataChanged.connect(self._on_data_changed)

        header = table.horizontalHeader()
        header.setContextMenuPolicy(Qt.CustomContextMenu)
        header.customContextMenuRequested.connect(self._on_header_context_menu)

        # 엑셀 레이아웃 먼저 적용
        self._apply_excel_layout(ws)
        QApplication.processEvents()
        
        # 내용에 맞게 컬럼 너비와 행 높이 자동 조정
        table.resizeColumnsToContents()
        QApplication.processEvents()
        table.resizeRowsToContents()
        QApplication.processEvents()
        
        # 컬럼 너비: 엑셀 원본보다 작아지지 않도록
        col_count = self.proxy.columnCount()
        for col_idx in range(col_count):
            if col_idx % 10 == 0:  # 10개 컬럼마다 UI 이벤트 처리
                QApplication.processEvents()
            current_width = table.columnWidth(col_idx)
            excel_col_name = ExcelSheetModel.excel_col_name(col_idx + 1)
            dim = ws.column_dimensions.get(excel_col_name)
            if dim and dim.width:
                excel_width = int(dim.width * 7 + 12)
                table.setColumnWidth(col_idx, max(current_width, excel_width))
        
        # 행 높이: 엑셀 원본보다 작아지지 않도록
        row_count = self.proxy.rowCount()
        # 행이 많을 수 있으므로 샘플링하거나 처리 속도 최적화
        if row_count < 1000:  # 행이 너무 많으면 생략하거나 최적화
            for row_idx in range(row_count):
                if row_idx % 50 == 0:
                    QApplication.processEvents()
                current_height = table.rowHeight(row_idx)
                dim = ws.row_dimensions.get(row_idx + 1)
                if dim and dim.height:
                    excel_height = int(dim.height * 1.33)
                    table.setRowHeight(row_idx, max(current_height, excel_height))

        self.on_search_changed(self.control_panel.get_search_edit().text())
        QApplication.processEvents()

    def on_sheet_changed(self, sheet_name: str):
        if self.model:
            self.model.apply_dirty_to_sheet()
        self.load_sheet(sheet_name)
        # 시트 변경 시 전처리 버튼 상태 업데이트
        self._update_preprocess_button_state()

    # ================= 검색 =================
    def on_search_changed(self, text: str):
        if not self.proxy:
            return
        if not text:
            self.proxy.setFilterRegularExpression(QRegularExpression(""))
            return
        rx = QRegularExpression(
            QRegularExpression.escape(text),
            QRegularExpression.CaseInsensitiveOption
        )
        self.proxy.setFilterRegularExpression(rx)

    # ================= 편집 모드 =================
    def on_edit_mode_changed(self):
        if self.model:
            edit_all = self.control_panel.get_edit_all_checkbox().isChecked()
            self.model.set_edit_all(edit_all)
            self.model.layoutChanged.emit()
    
    # ================= Undo/Redo =================
    def on_undo(self):
        """실행취소 버튼 클릭"""
        if self.model and self.model.undo():
            self._update_undo_redo_buttons()
    
    def on_redo(self):
        """다시실행 버튼 클릭"""
        if self.model and self.model.redo():
            self._update_undo_redo_buttons()
    
    def _update_undo_redo_buttons(self):
        """Undo/Redo 버튼 상태 업데이트"""
        if self.model:
            self.control_panel.get_undo_button().setEnabled(self.model.can_undo())
            self.control_panel.get_redo_button().setEnabled(self.model.can_redo())
        else:
            self.control_panel.get_undo_button().setEnabled(False)
            self.control_panel.get_redo_button().setEnabled(False)
    
    def _on_data_changed(self, top_left, bottom_right, roles):
        """데이터 변경 시 Undo/Redo 버튼 상태 업데이트"""
        self._update_undo_redo_buttons()

    # ================= 전처리 =================
    def on_preprocess_clicked(self):
        # 현재 선택된 시트가 있는지 확인
        sheet_combo = self.control_panel.get_sheet_combo()
        current_sheet = sheet_combo.currentText()
        if not current_sheet:
            QMessageBox.information(self, "안내", "먼저 파일을 업로드하세요.")
            return

        # 현재 시트의 워크북 찾기
        file_type = None
        if current_sheet.startswith("국내: "):
            if not self.wb_domestic:
                QMessageBox.information(self, "안내", "국내 청구서가 없습니다.")
                return
            wb = self.wb_domestic
            file_type = "domestic"
        elif current_sheet.startswith("해외: "):
            if not self.wb_overseas:
                QMessageBox.information(self, "안내", "해외 청구서가 없습니다.")
                return
            wb = self.wb_overseas
            file_type = "overseas"
        else:
            QMessageBox.information(self, "안내", "시트를 선택하세요.")
            return

        # 이미 전처리된 경우 확인
        if file_type == "domestic" and self.preprocessed_domestic:
            QMessageBox.information(self, "안내", "국내 청구서는 이미 전처리되었습니다.")
            return
        elif file_type == "overseas" and self.preprocessed_overseas:
            QMessageBox.information(self, "안내", "해외 청구서는 이미 전처리되었습니다.")
            return

        if self.model:
            self.model.apply_dirty_to_sheet()

        # 로딩 애니메이션 표시
        self.preview_container.show_loading("전처리 중")
        QApplication.processEvents()
        
        # 모델 잠시 해제 (백그라운드 작업 중 시트 접근 방지)
        if self.model:
            self.model = None
            self.preview_container.get_table().setModel(None)
        
        # 백그라운드에서 전처리 실행
        company = self.control_panel.get_company_edit().text().strip()
        keyword = self.control_panel.get_search_edit().text().strip()
        
        self.process_worker = WorkerThread(preprocess_inplace, wb, company=company, keyword=keyword)
        self.process_worker.finished.connect(lambda _: self._on_preprocess_finished(file_type, current_sheet))
        self.process_worker.error.connect(self._on_worker_error)
        self.process_worker.start()

    def _on_preprocess_finished(self, file_type, current_sheet):
        """전처리 완료 시 호출되는 콜백"""
        # 전처리 상태 업데이트
        if file_type == "domestic":
            self.preprocessed_domestic = True
        else:
            self.preprocessed_overseas = True

        # 전처리 버튼 상태 업데이트
        self._update_preprocess_button_state()

        self.info_panel.set_remark("전처리 완료. 미리보기 갱신됨")
        
        # 시트 다시 로드 (무거운 작업)
        self.load_sheet(current_sheet)
        
        # 모든 처리가 끝난 후 로딩 애니메이션 숨김
        QApplication.processEvents()
        self.preview_container.hide_loading()

    def _on_worker_error(self, message):
        """작업 도중 에러 발생 시 호출되는 콜백"""
        self.preview_container.hide_loading()
        QMessageBox.critical(self, "오류", message)
    
    def _update_preprocess_button_state(self):
        """전처리 버튼 상태 업데이트 (현재 선택된 시트에 따라)"""
        btn_preprocess = self.control_panel.get_preprocess_button()
        sheet_combo = self.control_panel.get_sheet_combo()
        current_sheet = sheet_combo.currentText()
        
        if not current_sheet:
            btn_preprocess.setText("전처리")
            btn_preprocess.setEnabled(False)
            return
        
        # 현재 시트 타입 확인
        if current_sheet.startswith("국내: "):
            if self.preprocessed_domestic:
                btn_preprocess.setText("전처리완료")
                btn_preprocess.setEnabled(False)
            else:
                btn_preprocess.setText("전처리")
                btn_preprocess.setEnabled(True)
        elif current_sheet.startswith("해외: "):
            if self.preprocessed_overseas:
                btn_preprocess.setText("전처리완료")
                btn_preprocess.setEnabled(False)
            else:
                btn_preprocess.setText("전처리")
                btn_preprocess.setEnabled(True)
        else:
            btn_preprocess.setText("전처리")
            btn_preprocess.setEnabled(True)

    # ================= 저장 =================
    def save_as_file(self):
        # 현재 선택된 시트의 dirty 데이터 저장
        if self.model:
            self.model.apply_dirty_to_sheet()

        # 저장할 워크북 결정
        if self.wb_domestic and self.wb_overseas:
            # 둘 다 있으면 합쳐서 저장
            from openpyxl import Workbook
            merged_wb = Workbook()
            merged_wb.remove(merged_wb.active)  # 기본 시트 제거
            
            # 국내 시트들 복사
            for sheet_name in self.wb_domestic.sheetnames:
                source_sheet = self.wb_domestic[sheet_name]
                new_sheet = merged_wb.create_sheet(f"국내_{sheet_name}")
                self._copy_sheet(source_sheet, new_sheet)
            
            # 해외 시트들 복사
            for sheet_name in self.wb_overseas.sheetnames:
                source_sheet = self.wb_overseas[sheet_name]
                new_sheet = merged_wb.create_sheet(f"해외_{sheet_name}")
                self._copy_sheet(source_sheet, new_sheet)
            
            wb_to_save = merged_wb
        elif self.wb_domestic:
            wb_to_save = self.wb_domestic
        elif self.wb_overseas:
            wb_to_save = self.wb_overseas
        else:
            QMessageBox.information(self, "안내", "먼저 파일을 업로드하세요.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "최종 엑셀로 저장", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            save_workbook_safe(wb_to_save, Path(path))
            QMessageBox.information(self, "완료", "저장했습니다.")
        except AppError as e:
            QMessageBox.critical(self, "오류", str(e))
    
    def _copy_sheet(self, source_sheet, target_sheet):
        """시트 내용 복사 (수식 포함)"""
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                
                # 수식이 있으면 수식 복사, 없으면 값 복사
                if cell.data_type == 'f':  # formula
                    target_cell.value = cell.value  # 수식 문자열
                else:
                    target_cell.value = cell.value
                
                # 스타일 복사
                if cell.has_style:
                    target_cell.font = cell.font
                    target_cell.border = cell.border
                    target_cell.fill = cell.fill
                    target_cell.number_format = cell.number_format
                    target_cell.protection = cell.protection
                    target_cell.alignment = cell.alignment
        
        # 병합 셀 복사
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
        
        # 열 너비 복사
        for col in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
        
        # 행 높이 복사
        for row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

    # ================= 테이블 헤더 메뉴 =================
    def _on_header_context_menu(self, pos):
        if not self.proxy or not self.model:
            return

        table = self.preview_container.get_table()
        header = table.horizontalHeader()
        col = header.logicalIndexAt(pos)
        if col < 0:
            return

        menu = QMenu(self)
        act_filter = menu.addAction("필터...")
        act_clear = menu.addAction("이 컬럼 필터 해제")
        act_clear_all = menu.addAction("전체 필터 초기화")

        picked = menu.exec(header.mapToGlobal(pos))
        if not picked:
            return

        if picked == act_filter:
            col_name = ExcelSheetModel.excel_col_name(col + 1)
            ColumnFilterDialog(self.model, self.proxy, col, col_name, self).exec()
        elif picked == act_clear:
            self.proxy.clear_column_filter(col)
            self._update_filter_button_state()
        elif picked == act_clear_all:
            self.proxy.clear_all_column_filters()
            self._update_filter_button_state()

    # ================= 필터 =================
    def on_filter_button_clicked(self):
        """필터 버튼 클릭 시 컬럼 선택 후 필터 다이얼로그 열기"""
        if not self.model or not self.proxy:
            QMessageBox.information(self, "안내", "먼저 파일을 업로드하세요.")
            return
        
        # 컬럼 선택 다이얼로그 열기
        col_dialog = ColumnSelectDialog(self.model, self)
        if col_dialog.exec() == QDialog.Accepted:
            col = col_dialog.get_selected_column()
            if col is not None:
                col_name = ExcelSheetModel.excel_col_name(col + 1)
                # 필터 다이얼로그 열기
                ColumnFilterDialog(self.model, self.proxy, col, col_name, self).exec()
                # 필터 상태 업데이트
                self._update_filter_button_state()
    
    def on_clear_filter_clicked(self):
        """필터 해제 버튼 클릭 시 모든 필터 해제"""
        if not self.proxy:
            return
        
        self.proxy.clear_all_column_filters()
        self._update_filter_button_state()
    
    def _update_filter_button_state(self):
        """필터 상태에 따라 필터 해제 버튼 활성화/비활성화"""
        if self.proxy and isinstance(self.proxy, ExcelFilterProxyModel):
            has_filters = self.proxy.has_active_filters()
            self.control_panel.get_clear_filter_button().setEnabled(has_filters)
        else:
            self.control_panel.get_clear_filter_button().setEnabled(False)
        
        # 필터 변경 후 병합 셀 다시 적용
        self._apply_merged_cells_only()
    
    def _apply_merged_cells_only(self):
        """병합 셀만 다시 적용 (필터 변경 후)"""
        if not self.model or not hasattr(self.model, 'ws'):
            return
        
        table = self.preview_container.get_table()
        table.clearSpans()
        
        ws = self.model.ws
        for mr in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = mr.bounds
            row = min_row - 1
            col = min_col - 1
            row_span = max_row - min_row + 1
            col_span = max_col - min_col + 1
            table.setSpan(row, col, row_span, col_span)
    
    # ================= 엑셀 레이아웃 =================
    def _apply_excel_layout(self, ws):
        table = self.preview_container.get_table()

        for col_idx in range(1, ws.max_column + 1):
            dim = ws.column_dimensions.get(
                ExcelSheetModel.excel_col_name(col_idx)
            )
            if dim and dim.width:
                table.setColumnWidth(col_idx - 1, int(dim.width * 7 + 12))

        for row_idx in range(1, ws.max_row + 1):
            dim = ws.row_dimensions.get(row_idx)
            if dim and dim.height:
                table.setRowHeight(row_idx - 1, int(dim.height * 1.33))
        
        # 병합 셀 처리: setSpan으로 병합 표시
        for mr in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = mr.bounds
            # QTableWidget의 인덱스는 0부터 시작
            row = min_row - 1
            col = min_col - 1
            row_span = max_row - min_row + 1
            col_span = max_col - min_col + 1
            table.setSpan(row, col, row_span, col_span)
