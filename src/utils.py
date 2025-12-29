"""
공통 유틸리티 함수들
"""
from datetime import datetime, date, timedelta


def norm(v) -> str:
    """문자열 정규화 (개행 제거, 소문자 변환)"""
    return str(v).replace("\n", " ").strip().lower()


def find_col_by_keywords_ws(ws, header_row: int, keywords: list[str], mode: str = "any") -> int:
    """
    ws에서 header_row를 기준으로 keywords로 컬럼 찾기(1-index)

    mode:
      - "any": keywords 중 하나라도 포함되면 매칭(OR)
      - "all": keywords 전부 포함되어야 매칭(AND)
    """
    mode = (mode or "any").lower()
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v in (None, ""):
            continue
        s = norm(v)

        if mode == "all":
            ok = all(k.lower() in s for k in keywords)
        else:
            ok = any(k.lower() in s for k in keywords)

        if ok:
            return col

    raise ValueError(f"컬럼을 찾을 수 없습니다: {keywords} (mode={mode})")


def parse_int_like(v):
    """주행거리 등 숫자 파싱(콤마/문자 섞여도 최대한)"""
    if v in (None, ""):
        return None
    try:
        if isinstance(v, (int, float)):
            return int(float(v))
        s = str(v).strip().replace(",", "")
        if not s:
            return None
        return int(float(s))
    except:
        return None


def parse_excel_date(v):
    """
    파일별 날짜 표기가 달라도 최대한 파싱:
    - datetime/date
    - 20250725 같은 int/str(yyyymmdd)
    - 2025-07-25, 2025/07/25, 2025.07.25 등
    - 엑셀 시리얼(대략 20000~80000 범위)도 처리
    """
    if v in (None, ""):
        return None

    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    # numeric: yyyymmdd or excel serial
    if isinstance(v, (int, float)):
        iv = int(v)

        # yyyymmdd 추정
        if 19000101 <= iv <= 21001231:
            s = str(iv)
            try:
                return datetime.strptime(s, "%Y%m%d").date()
            except:
                pass

        # excel serial 추정(대략)
        if 20000 <= iv <= 80000:
            base = datetime(1899, 12, 30)  # Excel 관행
            try:
                return (base + timedelta(days=float(v))).date()
            except:
                return None

    # string parse
    s = str(v).strip()
    if not s:
        return None

    # yyyymmdd 문자열(구분자 제거)
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d").date()
        except:
            pass

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y %m %d",
                "%y-%m-%d", "%y/%m/%d", "%y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            continue

    return None


def guess_last_data_row(ws, data_start_row: int, anchor_col: int, empty_run: int = 30) -> int:
    """
    데이터 끝 추정:
    anchor_col(예: repair_date)을 기준으로 연속 empty가 empty_run 이상이면 그 전을 데이터 끝으로 봄
    """
    last = ws.max_row
    streak = 0
    for r in range(data_start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=anchor_col).value
        if v in (None, ""):
            streak += 1
            if streak >= empty_run:
                last = r - empty_run
                break
        else:
            streak = 0
    return max(last, data_start_row)

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class AppError(Exception):
    """UI에서 사용자에게 메시지로 보여줄 목적의 예외"""
    pass


def load_workbook_safe(path: Path) -> Workbook:
    try:
        return load_workbook(path, data_only=False)
    except Exception as e:
        raise AppError(f"엑셀 로드 실패: {path}\n{e}") from e


def save_workbook_safe(wb: Workbook, path: Path) -> None:
    try:
        wb.save(path)
    except Exception as e:
        raise AppError(f"엑셀 저장 실패: {path}\n{e}") from e
