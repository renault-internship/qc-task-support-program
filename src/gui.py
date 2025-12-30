# gui.py
from __future__ import annotations

from pathlib import Path
from typing import Dict, Any

from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex
from PySide6.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QTableView,
    QPushButton, QFileDialog, QMessageBox, QComboBox,
    QLineEdit, QLabel, QGroupBox, QFormLayout,
    QDialog, QSpinBox, QDoubleSpinBox, QTableWidget, QTableWidgetItem
)

from openpyxl.workbook.workbook import Workbook

from src.utils import load_workbook_safe, save_workbook_safe, AppError
from src.excel_processor import preprocess_inplace
from src.database import (
    init_database, get_company_info, get_all_companies, 
    get_rules_from_table, add_rule_to_table
)
from typing import List


class AddRuleDialog(QDialog):
    """Rule 추가 다이얼로그"""
    def __init__(self, rule_table_name: str, parent=None):
        super().__init__(parent)
        self.rule_table_name = rule_table_name
        self.setWindowTitle(f"Rule 추가 - {rule_table_name}")
        self.setFixedSize(500, 600)
        
        layout = QFormLayout()
        
        # Priority
        self.priority_spin = QSpinBox()
        self.priority_spin.setRange(1, 999)
        self.priority_spin.setValue(100)
        layout.addRow("우선순위 *:", self.priority_spin)
        
        # Status
        self.status_combo = QComboBox()
        self.status_combo.addItems(["ACTIVE", "INACTIVE"])
        layout.addRow("상태 *:", self.status_combo)
        
        # Repair Region
        self.repair_region_edit = QLineEdit()
        layout.addRow("수리 지역 *:", self.repair_region_edit)
        
        # Vehicle Classification
        self.vehicle_class_edit = QLineEdit()
        layout.addRow("차량 분류 *:", self.vehicle_class_edit)
        
        # Project Code (선택)
        self.project_code_edit = QLineEdit()
        self.project_code_edit.setPlaceholderText("비워두면 모든 프로젝트")
        layout.addRow("프로젝트 코드:", self.project_code_edit)
        
        # Part Name (선택)
        self.part_name_edit = QLineEdit()
        self.part_name_edit.setPlaceholderText("비워두면 모든 부품")
        layout.addRow("부품명:", self.part_name_edit)
        
        # Part No (선택)
        self.part_no_edit = QLineEdit()
        self.part_no_edit.setPlaceholderText("선택사항")
        layout.addRow("부품 번호:", self.part_no_edit)
        
        # Liability Ratio (필수)
        self.liability_ratio_spin = QDoubleSpinBox()
        self.liability_ratio_spin.setRange(0.0, 100.0)
        self.liability_ratio_spin.setDecimals(2)
        self.liability_ratio_spin.setSuffix(" %")
        self.liability_ratio_spin.setValue(0.0)
        layout.addRow("구상율 *:", self.liability_ratio_spin)
        
        # Amount Cap Type
        self.amount_cap_combo = QComboBox()
        self.amount_cap_combo.addItems(["NONE", "FIXED", "PERCENTAGE"])
        layout.addRow("금액 상한 타입 *:", self.amount_cap_combo)
        
        # Amount Cap Value
        self.amount_cap_spin = QSpinBox()
        self.amount_cap_spin.setRange(0, 999999999)
        self.amount_cap_spin.setValue(0)
        layout.addRow("금액 상한 값:", self.amount_cap_spin)
        
        # Exclude Project Code
        self.exclude_project_code_edit = QLineEdit()
        self.exclude_project_code_edit.setPlaceholderText("제외할 프로젝트 코드")
        layout.addRow("제외 프로젝트 코드:", self.exclude_project_code_edit)
        
        # Note
        self.note_edit = QLineEdit()
        layout.addRow("비고:", self.note_edit)
        
        # 버튼
        button_layout = QHBoxLayout()
        self.save_btn = QPushButton("저장")
        self.save_btn.clicked.connect(self.accept)
        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addRow(button_layout)
        
        self.setLayout(layout)
    
    def get_data(self):
        """입력된 데이터 반환"""
        return {
            "priority": self.priority_spin.value(),
            "status": self.status_combo.currentText(),
            "repair_region": self.repair_region_edit.text().strip(),
            "vehicle_classification": self.vehicle_class_edit.text().strip(),
            "project_code": self.project_code_edit.text().strip(),
            "part_name": self.part_name_edit.text().strip(),
            "part_no": self.part_no_edit.text().strip(),
            "liability_ratio": self.liability_ratio_spin.value(),
            "amount_cap_type": self.amount_cap_combo.currentText(),
            "amount_cap_value": self.amount_cap_spin.value() if self.amount_cap_spin.value() > 0 else None,
            "exclude_project_code": self.exclude_project_code_edit.text().strip(),
            "note": self.note_edit.text().strip(),
        }


class ViewRulesDialog(QDialog):
    """Rule 목록 보기 다이얼로그 (변경점만 표시)"""
    def __init__(self, rules: List[Dict[str, Any]], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Rule 목록")
        self.setFixedSize(800, 500)
        
        layout = QVBoxLayout()
        
        # Rule 목록 테이블
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["우선순위", "상태", "변경점"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setAlternatingRowColors(True)
        
        # Rule 데이터 채우기
        self.populate_rules(rules)
        
        layout.addWidget(self.table)
        
        # 닫기 버튼
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        close_btn = QPushButton("닫기")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def format_rule_changes(self, rule: Dict[str, Any]) -> str:
        """Rule의 변경점만 포맷팅하여 반환"""
        changes = []
        
        # 프로젝트 코드
        project_code = rule.get("project_code", "").strip()
        if project_code:
            changes.append(f"프로젝트: {project_code}")
        
        # 부품명
        part_name = rule.get("part_name", "").strip()
        if part_name:
            changes.append(f"부품: {part_name}")
        
        # 부품 번호
        part_no = rule.get("part_no", "").strip()
        if part_no:
            changes.append(f"부품번호: {part_no}")
        
        # 제외 프로젝트
        exclude_project = rule.get("exclude_project_code", "").strip()
        if exclude_project:
            changes.append(f"제외: {exclude_project}")
        
        # 구상율 (항상 표시)
        liability_ratio = rule.get("liability_ratio", 0)
        if liability_ratio is not None:
            changes.append(f"구상율: {liability_ratio}%")
        
        # 보증 주행거리 오버라이드
        warranty_mileage = rule.get("warranty_mileage_override")
        if warranty_mileage:
            changes.append(f"주행거리: {warranty_mileage}km")
        
        # 보증 기간 오버라이드
        warranty_period = rule.get("warranty_period_override")
        if warranty_period:
            years = warranty_period / 365.0
            changes.append(f"보증기간: {years:.1f}년")
        
        # 금액 상한
        amount_cap_value = rule.get("amount_cap_value")
        if amount_cap_value:
            cap_type = rule.get("amount_cap_type", "NONE")
            if cap_type != "NONE":
                changes.append(f"상한: {amount_cap_value} ({cap_type})")
        
        # 비고
        note = rule.get("note", "").strip()
        if note:
            changes.append(f"비고: {note}")
        
        return " | ".join(changes) if changes else "기본 규칙"
    
    def populate_rules(self, rules: List[Dict[str, Any]]):
        """Rule 목록을 테이블에 채우기"""
        self.table.setRowCount(len(rules))
        
        for row, rule in enumerate(rules):
            # 우선순위
            priority_item = QTableWidgetItem(str(rule.get("priority", "")))
            priority_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 0, priority_item)
            
            # 상태
            status = rule.get("status", "")
            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)
            # ACTIVE는 초록색, INACTIVE는 회색으로 표시
            if status.upper() == "ACTIVE":
                status_item.setForeground(Qt.GlobalColor.green)
            elif status.upper() == "INACTIVE":
                status_item.setForeground(Qt.GlobalColor.gray)
            self.table.setItem(row, 1, status_item)
            
            # 변경점
            changes_item = QTableWidgetItem(self.format_rule_changes(rule))
            self.table.setItem(row, 2, changes_item)
        
        self.table.resizeColumnsToContents()


class MainWindow(QWidget):
    """
    - 좌측: 미리보기(QTableView) + 시트 선택 + 정보 패널
    - 우측: 업로드/전처리/기업선택/search/export
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle("엑셀 전처리 도구")

        self.file_path: Path | None = None
        self.wb: Workbook | None = None
        self.model: ExcelSheetModel | None = None
        self.current_company_info: Dict[str, Any] | None = None
        
        # 데이터베이스 초기화
        init_database()

        # ===== 좌측: 미리보기 =====
        self.sheet_combo = QComboBox()
        self.sheet_combo.currentTextChanged.connect(self.on_sheet_changed)

        self.table = QTableView()
        self.table.setAlternatingRowColors(True)

        left_top = QHBoxLayout()
        left_top.addWidget(QLabel("시트"))
        left_top.addWidget(self.sheet_combo, 1)

        left_preview_box = QVBoxLayout()
        left_preview_box.addLayout(left_top)
        left_preview_box.addWidget(self.table, 1)

        info_group = QGroupBox("정보")
        form = QFormLayout()

        self.lbl_company = QLabel("-")
        self.lbl_remark = QLabel("-")
        self.lbl_editable = QLabel("-")
        
        # Rule 추가 버튼
        self.btn_add_rule = QPushButton("+ Rule 추가")
        self.btn_add_rule.setToolTip("Rule 추가")
        self.btn_add_rule.clicked.connect(self.add_rule)

        for lbl in (self.lbl_company, self.lbl_remark, self.lbl_editable):
            lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        
        # lbl_editable 클릭 시 Rule 목록 보기
        self.lbl_editable.mousePressEvent = self.show_rules_dialog
        self.lbl_editable.setCursor(Qt.CursorShape.PointingHandCursor)
        self.lbl_editable.setStyleSheet("QLabel { color: blue; text-decoration: underline; }")

        form.addRow("기업명", self.lbl_company)
        form.addRow("비고(remark)", self.lbl_remark)
        form.addRow("변경가능(rule)", self.lbl_editable)
        form.addRow("", self.btn_add_rule)
        info_group.setLayout(form)

        left_preview_box.addWidget(info_group)

        left = QWidget()
        left.setLayout(left_preview_box)

        # ===== 우측: 컨트롤 =====
        self.btn_upload = QPushButton("업로드")
        self.btn_preprocess = QPushButton("전처리")
        
        # 기업 선택
        company_label = QLabel("기업 선택")
        self.company_combo = QComboBox()
        self.load_companies()  # DB에서 기업 목록 로드
        self.company_combo.currentTextChanged.connect(self._on_company_changed)
        
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("search")

        self.btn_export_rule = QPushButton("export (rule)")
        self.btn_export_final = QPushButton("export (최종 엑셀)")

        self.btn_upload.clicked.connect(self.open_file)
        self.btn_preprocess.clicked.connect(self.on_preprocess_clicked)
        self.btn_export_rule.clicked.connect(self.export_rule_stub)
        self.btn_export_final.clicked.connect(self.save_as_file)

        right_box = QVBoxLayout()
        right_box.addWidget(self.btn_upload)
        right_box.addWidget(self.btn_preprocess)
        right_box.addSpacing(8)
        right_box.addWidget(company_label)
        right_box.addWidget(self.company_combo)
        right_box.addSpacing(8)
        right_box.addWidget(self.search_edit)
        right_box.addStretch(1)
        right_box.addWidget(self.btn_export_rule)
        right_box.addWidget(self.btn_export_final)

        right = QWidget()
        right.setFixedWidth(220)
        right.setLayout(right_box)

        root = QHBoxLayout()
        root.addWidget(left, 1)
        root.addWidget(right)
        self.setLayout(root)

        self._set_info_defaults()

    def _set_info_defaults(self):
        self.lbl_company.setText("-")
        self.lbl_remark.setText("-")
        self.lbl_editable.setText("-")
    
    def load_companies(self):
        """기업 목록 로드 (DB에서)"""
        self.company_combo.clear()
        companies = get_all_companies()
        if companies:
            self.company_combo.addItem("선택")
            self.company_combo.addItems(companies)
        else:
            self.company_combo.addItem("선택")
    
    def _on_company_changed(self, name: str):
        """기업 선택 변경 시 정보 업데이트"""
        if name and name != "선택":
            # sap 테이블에서 기업 정보 가져오기
            company_info = get_company_info(name)
            if company_info:
                self.lbl_company.setText(name)
                
                # rule_table_name 가져오기
                rule_table_name = company_info.get("rule_table_name")
                if rule_table_name:
                    # rule 테이블에서 rule 정보 조회
                    rules = get_rules_from_table(rule_table_name)
                    if rules:
                        # rule 개수와 주요 정보 표시
                        rule_count = len(rules)
                        active_rules = [r for r in rules if r.get("status", "").upper() == "ACTIVE"]
                        active_count = len(active_rules)
                        self.lbl_editable.setText(f"Rule: {rule_count}개 (활성: {active_count}개)")
                    else:
                        self.lbl_editable.setText(f"Rule 테이블: {rule_table_name} (규칙 없음)")
                else:
                    self.lbl_editable.setText("Rule 테이블 없음")
                
                # 현재 선택된 기업 정보 저장 (rule 추가 시 사용)
                self.current_company_info = company_info
            else:
                self.lbl_company.setText("-")
                self.lbl_editable.setText("-")
                self.current_company_info = None
        else:
            self.lbl_company.setText("-")
            self.lbl_editable.setText("-")
            self.current_company_info = None
    
    def add_rule(self):
        """Rule 추가 다이얼로그 열기"""
        # 기업이 선택되어 있는지 확인
        company_name = self.company_combo.currentText()
        if company_name == "선택" or not company_name:
            QMessageBox.warning(self, "오류", "먼저 기업을 선택해주세요.")
            return
        
        # 기업 정보 가져오기
        if not hasattr(self, 'current_company_info') or not self.current_company_info:
            company_info = get_company_info(company_name)
            if not company_info:
                QMessageBox.warning(self, "오류", f"기업정보를 찾을 수 없습니다: {company_name}")
                return
            self.current_company_info = company_info
        
        rule_table_name = self.current_company_info.get("rule_table_name")
        if not rule_table_name:
            QMessageBox.warning(self, "오류", "선택한 기업에 Rule 테이블이 없습니다.")
            return
        
        # Rule 추가 다이얼로그 열기
        dialog = AddRuleDialog(rule_table_name, self)
        if dialog.exec() == QDialog.Accepted:
            data = dialog.get_data()
            
            # 필수 필드 검증
            if not data["repair_region"]:
                QMessageBox.warning(self, "오류", "수리 지역을 입력해주세요.")
                return
            if not data["vehicle_classification"]:
                QMessageBox.warning(self, "오류", "차량 분류를 입력해주세요.")
                return
            
            try:
                rule_id = add_rule_to_table(
                    rule_table_name=rule_table_name,
                    priority=data["priority"],
                    status=data["status"],
                    repair_region=data["repair_region"],
                    vehicle_classification=data["vehicle_classification"],
                    liability_ratio=data["liability_ratio"],
                    amount_cap_type=data["amount_cap_type"],
                    project_code=data["project_code"],
                    part_name=data["part_name"],
                    part_no=data["part_no"],
                    exclude_project_code=data["exclude_project_code"],
                    amount_cap_value=data["amount_cap_value"],
                    note=data["note"],
                )
                
                QMessageBox.information(self, "완료", f"Rule이 추가되었습니다. (ID: {rule_id})")
                
                # 정보 패널 업데이트 (rule 개수 갱신)
                self._on_company_changed(company_name)
            except Exception as e:
                QMessageBox.critical(self, "오류", f"Rule 추가 실패: {str(e)}")
    
    def show_rules_dialog(self, event):
        """Rule 목록 다이얼로그 표시"""
        company_name = self.company_combo.currentText()
        if company_name == "선택" or not company_name:
            QMessageBox.information(self, "안내", "먼저 기업을 선택해주세요.")
            return
        
        company_info = get_company_info(company_name)
        if not company_info:
            QMessageBox.warning(self, "오류", f"기업정보를 찾을 수 없습니다: {company_name}")
            return
        
        rule_table_name = company_info.get("rule_table_name")
        if not rule_table_name:
            QMessageBox.information(self, "안내", "선택한 기업에 Rule 테이블이 없습니다.")
            return
        
        rules = get_rules_from_table(rule_table_name)
        if not rules:
            QMessageBox.information(self, "안내", "등록된 Rule이 없습니다.")
            return
        
        dialog = ViewRulesDialog(rules, self)
        dialog.exec()

    # ---------- 업로드 ----------
    def open_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "엑셀 선택", "", "Excel Files (*.xlsx)")
        if not path:
            return

        self.file_path = Path(path)

        try:
            self.wb = load_workbook_safe(self.file_path)
        except AppError as e:
            QMessageBox.critical(self, "오류", str(e))
            return

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        self.sheet_combo.addItems(self.wb.sheetnames)
        self.sheet_combo.blockSignals(False)

        if self.wb.sheetnames:
            self.sheet_combo.setCurrentIndex(0)
            self.load_sheet(self.wb.sheetnames[0])

        company = self.company_combo.currentText()
        self.lbl_company.setText(company if company != "선택" else "-")
        self.lbl_remark.setText("업로드 완료. 전처리 전 상태")
        self.lbl_editable.setText("현재: 전체 셀 편집 가능(추후 rule로 제한)")

    # ---------- 시트 로드/변경 ----------
    def load_sheet(self, sheet_name: str):
        if not self.wb:
            return
        ws = self.wb[sheet_name]
        self.model = ExcelSheetModel(ws, self)
        self.table.setModel(self.model)
        self.table.resizeColumnsToContents()

    def on_sheet_changed(self, sheet_name: str):
        if self.model:
            self.model.apply_dirty_to_sheet()
        self.load_sheet(sheet_name)

    # ---------- 전처리 ----------
    def on_preprocess_clicked(self):
        if not self.wb:
            QMessageBox.information(self, "안내", "먼저 파일을 업로드하세요.")
            return

        # 미리보기에서 수정해둔 내용이 있으면 먼저 workbook에 반영
        if self.model:
            self.model.apply_dirty_to_sheet()

        company = self.company_combo.currentText()
        keyword = self.search_edit.text().strip()

        try:
            preprocess_inplace(self.wb, company=company, keyword=keyword)
        except AppError as e:
            QMessageBox.critical(self, "오류", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "오류", f"전처리 실패:\n{e}")
            return

        self.lbl_company.setText(company if company != "선택" else "-")
        self.lbl_remark.setText("전처리 완료. 미리보기 갱신됨")
        self.lbl_editable.setText("전처리 후: 필요한 경우 rule 기반 편집 제한 적용 예정")

        self.refresh_preview_after_processing()

    def refresh_preview_after_processing(self):
        if not self.wb:
            return

        current_sheet = self.sheet_combo.currentText()
        if not current_sheet or current_sheet not in self.wb.sheetnames:
            current_sheet = self.wb.sheetnames[0] if self.wb.sheetnames else ""

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        self.sheet_combo.addItems(self.wb.sheetnames)
        if current_sheet:
            self.sheet_combo.setCurrentText(current_sheet)
        self.sheet_combo.blockSignals(False)

        if current_sheet:
            self.load_sheet(current_sheet)

    # ---------- export ----------
    def save_as_file(self):
        if not self.wb:
            QMessageBox.information(self, "안내", "먼저 파일을 업로드하세요.")
            return

        if self.model:
            self.model.apply_dirty_to_sheet()

        save_path, _ = QFileDialog.getSaveFileName(self, "최종 엑셀로 저장", "", "Excel Files (*.xlsx)")
        if not save_path:
            return

        try:
            save_workbook_safe(self.wb, Path(save_path))
            QMessageBox.information(self, "완료", "저장했습니다.")
        except AppError as e:
            QMessageBox.critical(self, "오류", str(e))

    def export_rule_stub(self):
        QMessageBox.information(self, "안내", "rule export는 아직 연결 안 함.")


class ExcelSheetModel(QAbstractTableModel):
    def __init__(self, ws, parent=None):
        super().__init__(parent)
        self.ws = ws
        self.max_row = ws.max_row
        self.max_col = ws.max_column
        self.dirty = {}

    def rowCount(self, parent=QModelIndex()):
        return self.max_row

    def columnCount(self, parent=QModelIndex()):
        return self.max_col

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        r = index.row() + 1
        c = index.column() + 1
        if role in (Qt.DisplayRole, Qt.EditRole):
            v = self.dirty.get((r, c), self.ws.cell(row=r, column=c).value)
            return "" if v is None else str(v)
        return None

    def flags(self, index):
        if not index.isValid():
            return Qt.NoItemFlags
        return Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable

    def setData(self, index, value, role=Qt.EditRole):
        if role != Qt.EditRole or not index.isValid():
            return False
        r = index.row() + 1
        c = index.column() + 1
        text = (value or "").strip()

        new_val = None
        if text != "":
            raw = text.replace(",", "")
            try:
                if "." in raw:
                    new_val = float(raw)
                else:
                    new_val = int(raw)
            except ValueError:
                new_val = text

        self.dirty[(r, c)] = new_val
        self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
        return True

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return self._excel_col_name(section + 1)
        return str(section + 1)

    @staticmethod
    def _excel_col_name(n: int) -> str:
        name = ""
        while n:
            n, rem = divmod(n - 1, 26)
            name = chr(65 + rem) + name
        return name

    def apply_dirty_to_sheet(self):
        for (r, c), v in self.dirty.items():
            self.ws.cell(row=r, column=c).value = v
