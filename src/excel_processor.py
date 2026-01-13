

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, List, Tuple, Dict, Optional
from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import MergedCell

from src.utils import (
    find_col_by_keywords_ws,
    parse_int_like,
    parse_excel_date,
    guess_last_data_row,
    AppError,
)
from src.database import (
    get_common_project_liability,
    get_rules_from_table,
    _get_global_warranty,
)

FILL_HIGHLIGHT = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 전처리 1회 고정용 메타 시트
META_SHEET_NAME = "_PREPROCESS_META"
META_DONE_CELL = "A1"
META_TS_CELL = "A2"


# =========================================================
# 룰 설명 포맷팅 함수
# =========================================================
def format_rule_description(rule: dict) -> str:
    """룰의 변경점을 포맷팅하여 설명 반환"""
    changes = []
    
    def valid(val, ignore=("ALL", "NONE")):
        return val and str(val).strip().upper() not in ignore
    
    # 수리 지역 (ALL이 아닐 때만)
    if valid(rule.get("repair_region")):
        changes.append(f"수리지역:{rule['repair_region']}")
    
    # 프로젝트 코드 (ALL이 아닐 때만)
    if valid(rule.get("project_code")):
        changes.append(f"프로젝트:{rule['project_code']}")
    
    # 제외 프로젝트
    if rule.get("exclude_project_code"):
        changes.append(f"제외:{rule['exclude_project_code']}")
    
    # 차계 (ALL이 아닐 때만)
    if valid(rule.get("vehicle_classification")):
        changes.append(f"차계:{rule['vehicle_classification']}")
    
    # 부품명 (ALL이 아닐 때만)
    if valid(rule.get("part_name")):
        changes.append(f"부품:{rule['part_name']}")
    
    # 부품 번호 (ALL이 아닐 때만)
    if valid(rule.get("part_no")):
        changes.append(f"부품번호:{rule['part_no']}")
    
    # 엔진 형식 (ALL이 아닐 때만)
    if valid(rule.get("engine_form")):
        changes.append(f"엔진:{rule['engine_form']}")
    
    # 구상율 (항상 표시)
    if rule.get("liability_ratio") is not None:
        changes.append(f"구상율:{rule['liability_ratio'] * 100:.0f}%")
    
    # 보증 주행거리 오버라이드
    if rule.get("warranty_mileage_override") is not None:
        changes.append(f"주행거리:{rule['warranty_mileage_override']}km")
    
    # 보증 기간 오버라이드
    if rule.get("warranty_period_override") is not None:
        years = rule["warranty_period_override"] / 365.0
        changes.append(f"보증기간:{years:.1f}년")
    
    # 금액 상한
    if rule.get("amount_cap_value") is not None and valid(rule.get("amount_cap_type")):
        changes.append(f"상한:{rule['amount_cap_value']}({rule['amount_cap_type']})")
    
    # 적용 시작일
    if rule.get("valid_from"):
        changes.append(f"시작:{rule['valid_from']}")
    
    # 적용 종료일
    if rule.get("valid_to"):
        changes.append(f"종료:{rule['valid_to']}")
    
    return " | ".join(changes) if changes else "기본 규칙"


# =========================================================
# 전처리 결과 통계
# =========================================================
@dataclass
class PreprocessResult:
    """전처리 결과 통계"""
    # 기본 정보
    repair_region: str = ""  # DOMESTIC or OVERSEAS
    company_code: str = ""
    company_name: str = ""
    rule_table_name: str = ""
    process_time: str = ""  # 처리 일시
    
    # 처리 결과
    total_rows: int = 0
    success_rows: int = 0
    warning_rows: int = 0
    error_rows: int = 0
    
    # 차계 및 프로젝트 통계 {vehicle: (project_code, count, liability_ratio)}
    vehicle_stats: Dict[str, Tuple[str, int, Optional[float]]] = field(default_factory=dict)
    
    # 프로젝트 코드별 통계 {project_code: (count, liability_ratio)}
    project_stats: Dict[str, Tuple[int, Optional[float]]] = field(default_factory=dict)
    
    # 기본 구상률 적용 통계
    common_liability_applied: int = 0
    
    # 룰별 적용 통계 {rule_id: (description, count)}
    rule_usage: Dict[int, Tuple[str, int]] = field(default_factory=dict)
    
    # 미사용 룰 목록 [(rule_id, description, reason)]
    unused_rules: List[Tuple[int, str, str]] = field(default_factory=list)
    
    # 세부 룰 적용 통계
    warranty_mileage_rules_applied: int = 0
    warranty_period_rules_applied: int = 0
    liability_ratio_rules_applied: int = 0
    amount_cap_rules_applied: int = 0  # 공임비 상한 룰 적용 횟수
    amount_cap_exceeded_rows: int = 0  # 공임비 상한 초과 행 수
    
    # 보증 기준 (기본값 및 오버라이드)
    default_mileage_threshold: int = 0
    default_warranty_years: int = 0
    mileage_overrides: Dict[int, int] = field(default_factory=dict)  # {mileage_value: count}
    period_overrides: Dict[int, int] = field(default_factory=dict)  # {years: count}
    
    # 워런티 초과 통계
    mileage_exceeded_rows: int = 0  # 주행거리 초과
    period_exceeded_rows: int = 0   # 보증기간 초과
    both_exceeded_rows: int = 0     # 둘 다 초과
    warranty_highlighted_rows: int = 0  # 총 하이라이트
    
    # 경고 항목 (row_num, vehicle, part_no, reason)
    warnings: List[Tuple[int, str, str, str]] = field(default_factory=list)
    
    # 비고/로그 (사용자 확인 필요 항목)
    remarks: List[str] = field(default_factory=list)
    
    def add_vehicle_stat(self, vehicle: str, project_code: str, liability_ratio: Optional[float] = None):
        """차계별 통계 추가"""
        if vehicle in self.vehicle_stats:
            pc, count, ratio = self.vehicle_stats[vehicle]
            self.vehicle_stats[vehicle] = (pc, count + 1, ratio or liability_ratio)
        else:
            self.vehicle_stats[vehicle] = (project_code, 1, liability_ratio)
    
    def add_project_stat(self, project_code: str, liability_ratio: Optional[float] = None):
        """프로젝트 코드별 통계 추가"""
        if project_code in self.project_stats:
            count, ratio = self.project_stats[project_code]
            self.project_stats[project_code] = (count + 1, ratio or liability_ratio)
        else:
            self.project_stats[project_code] = (1, liability_ratio)
    
    def add_rule_usage(self, rule_id: int, description: str):
        """룰 사용 통계 추가"""
        if rule_id in self.rule_usage:
            desc, count = self.rule_usage[rule_id]
            self.rule_usage[rule_id] = (desc, count + 1)
        else:
            self.rule_usage[rule_id] = (description, 1)
    
    def add_mileage_override(self, mileage: int):
        """주행거리 오버라이드 통계 추가"""
        self.mileage_overrides[mileage] = self.mileage_overrides.get(mileage, 0) + 1
    
    def add_period_override(self, years: int):
        """보증기간 오버라이드 통계 추가"""
        self.period_overrides[years] = self.period_overrides.get(years, 0) + 1
    
    def add_warning(self, row_num: int, vehicle: str, part_no: str, reason: str):
        """경고 항목 추가"""
        self.warnings.append((row_num, vehicle, part_no, reason))
        self.warning_rows += 1
    
    def add_remark(self, message: str):
        """비고/로그 추가"""
        self.remarks.append(message)


# =========================================================
# 차계 → 프로젝트 코드 매핑
# =========================================================
def get_project_code_from_vehicle(vehicle: str) -> str:
    """
    차계에서 프로젝트 코드 추출
    LFD(G___), HZG(H___), LJL(J___), AR1(K___)
    """
    if not vehicle or not isinstance(vehicle, str):
        return "UNKNOWN"
    
    vehicle_upper = vehicle.strip().upper()
    if not vehicle_upper:
        return "UNKNOWN"
    
    first_char = vehicle_upper[0]
    
    project_mapping = {
        'G': 'LFD',
        'H': 'HZG',
        'J': 'LJL',
        'K': 'AR1',
    }
    
    return project_mapping.get(first_char, "UNKNOWN")


# =========================================================
# 0) 전처리 1회만 가능: 마킹/체크
# =========================================================
def _is_already_preprocessed(wb: Workbook) -> bool:
    if META_SHEET_NAME not in wb.sheetnames:
        return False
    ws = wb[META_SHEET_NAME]
    v = ws[META_DONE_CELL].value
    return str(v).strip() == "1"


def _mark_preprocessed(wb: Workbook) -> None:
    if META_SHEET_NAME in wb.sheetnames:
        ws = wb[META_SHEET_NAME]
    else:
        ws = wb.create_sheet(META_SHEET_NAME)
        ws.sheet_state = "hidden"  # 숨김 처리

    ws[META_DONE_CELL].value = "1"
    ws[META_TS_CELL].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# =========================================================
# 0-1) 시트별 국내 해외 구분후 처리
# =========================================================

def _detect_repair_region_by_sheetname(ws) -> str | None:
    """
    시트 이름에 '국내' or '해외'가 포함되면 해당 region 반환.
    아니면 None.
    """
    name = (ws.title or "").strip().lower()
    if "국내" in name:
        return "DOMESTIC"
    if "해외" in name:
        return "OVERSEAS"
    return None


def _infer_region_with_opposite(wb, idx: int) -> str | None:
    """
    1) 내 시트명에 국내/해외 있으면 그걸로 확정
    2) 없으면, 워크북 내 다른 시트들 중 국내/해외가 확정된 게 있으면 '반대'로 추정
    3) 둘 다 못하면 None
    """
    ws = wb.worksheets[idx]

    mine = _detect_repair_region_by_sheetname(ws)
    if mine:
        return mine

    other_known = None
    for j, other_ws in enumerate(wb.worksheets):
        if j == idx:
            continue
        r = _detect_repair_region_by_sheetname(other_ws)
        if r:
            other_known = r
            break

    if not other_known:
        return None

    return "OVERSEAS" if other_known == "DOMESTIC" else "DOMESTIC"

# =========================================================
# 1) 병합셀(MergedCell) 안전 처리
# =========================================================
def _resolve_merged_anchor(ws, row: int, col: int) -> Tuple[int, int]:
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return row, col

    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col

    return row, col


def _cell_safe(ws, row: int, col: int):
    ar, ac = _resolve_merged_anchor(ws, row, col)
    return ws.cell(row=ar, column=ac)


def set_cell_value_safe(ws, row: int, col: int, value: Any) -> None:
    _cell_safe(ws, row, col).value = value


def set_cell_fill_safe(ws, row: int, col: int, fill: PatternFill) -> None:
    _cell_safe(ws, row, col).fill = fill


# =========================================================
# 2) 기본 유틸
# =========================================================
def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def iter_data_rows(ws, data_start_row: int, last_row: int, anchor_col: int) -> List[int]:
    """
    anchor_col(보통 수리일자/클레임번호 등)에 값이 있는 행만 데이터 행으로 본다.
    국내 원본처럼 빈행+병합이 많은 경우 필수.
    """
    rows: List[int] = []
    for r in range(data_start_row, last_row + 1):
        v = ws.cell(row=r, column=anchor_col).value
        if not _is_blank(v):
            rows.append(r)
    return rows


# =========================================================
# 3) 컬럼 찾기(정확도 중요)
# =========================================================
def find_rate_col(ws, header_row: int) -> int:
    """
    구상율(Liability Ratio) 컬럼 정확히 찾기
    """
    try:
        return find_col_by_keywords_ws(ws, header_row, ["liability", "구상율"], mode="all")
    except Exception:
        return find_col_by_keywords_ws(ws, header_row, ["구상율"], mode="any")


def find_chargeback_col(ws, header_row: int) -> int:
    """
    구상금액(Chargeback Amount) 컬럼 정확히 찾기
    - '구상'만 넣으면 구상율에도 걸릴 수 있으므로 금액을 반드시 포함
    """
    try:
        return find_col_by_keywords_ws(ws, header_row, ["chargeback", "구상금액"], mode="all")
    except Exception:
        pass

    try:
        return find_col_by_keywords_ws(ws, header_row, ["구상금액"], mode="any")
    except Exception:
        pass

    return find_col_by_keywords_ws(ws, header_row, ["구상", "금액"], mode="all")


def pick_mileage_col(ws, header_row: int) -> int:
    """
    주행거리 컬럼 찾기 (KM 단위 컬럼만)
    - "Mileage Km" 또는 "주행거리 Km" 포함 컬럼만 찾음
    """
    return find_col_by_keywords_ws(ws, header_row, ["mileage km", "주행거리 km"], mode="any")


def find_130_percent_col(ws, header_row: int) -> Optional[int]:
    """
    130% 컬럼 찾기 (부품비의 1.3배)
    """
    try:
        return find_col_by_keywords_ws(ws, header_row, ["130%", "1.3"], mode="any")
    except Exception:
        return None


def find_labor_cost_col(ws, header_row: int) -> Optional[int]:
    """
    Labor Cost (공임대) 컬럼 찾기
    """
    try:
        return find_col_by_keywords_ws(ws, header_row, ["labor cost", "공임대"], mode="any")
    except Exception:
        return None


def find_outsource_labor_cost_col(ws, header_row: int) -> Optional[int]:
    """
    Labor Cost_Outsourcing (외주공임) 컬럼 찾기
    """
    try:
        return find_col_by_keywords_ws(ws, header_row, ["labor cost_outsourcing", "외주공임"], mode="any")
    except Exception:
        return None


def find_total_cost_col(ws, header_row: int) -> Optional[int]:
    """
    Total Cost (발생금액) 컬럼 찾기
    """
    try:
        return find_col_by_keywords_ws(ws, header_row, ["total cost", "발생금액"], mode="any")
    except Exception:
        return None


# =========================================================
# 4) 차계 병합 해제 + 채우기 (데이터 범위까지만)
# =========================================================
def unmerge_and_fill_column(ws, target_col: int, data_start_row: int, last_row: int) -> None:
    merged_ranges = list(ws.merged_cells.ranges)
    unmerged_rows = set()  # 병합셀 해제된 행 추적

    # 병합셀 해제 및 채우기
    for mr in merged_ranges:
        if (mr.min_col <= target_col <= mr.max_col) and (mr.min_row >= data_start_row):
            top_left = ws.cell(mr.min_row, mr.min_col).value
            ws.unmerge_cells(str(mr))
            # 병합셀 해제된 행 범위 기록
            for r in range(mr.min_row, min(mr.max_row, last_row) + 1):
                unmerged_rows.add(r)
                set_cell_value_safe(ws, r, target_col, top_left)

    # 병합셀이 해제된 범위 내에서만 빈칸 채우기
    if unmerged_rows:
        prev = None
        for r in range(data_start_row, last_row + 1):
            cur = ws.cell(row=r, column=target_col).value
            if r in unmerged_rows:  # 병합셀 해제된 행만 처리
                if _is_blank(cur):
                    if not _is_blank(prev):
                        set_cell_value_safe(ws, r, target_col, prev)
                else:
                    prev = cur
            else:
                # 병합셀 해제되지 않은 행은 prev 업데이트만 (빈칸 채우지 않음)
                if not _is_blank(cur):
                    prev = cur


# =========================================================
# 5) 구상율 변경(단일 진입점) + 바뀐 행 추적
# =========================================================
def set_rate(ws, row: int, rate_col: int, new_rate: float, changed_rows: set[int]) -> None:
    cell = _cell_safe(ws, row, rate_col)
    old = cell.value

    try:
        old_f = float(str(old).replace(",", "")) if not _is_blank(old) else None
    except Exception:
        old_f = None

    # DB의 소수값(0.6)을 퍼센트(60.0)로 변환하여 엑셀에 저장
    new_rate_percent = float(new_rate) * 100

    if old_f != new_rate_percent:
        cell.value = new_rate_percent
        changed_rows.add(row)


# =========================================================
# 6) 발생금액 수식(데이터 행만)
# =========================================================
def set_total_cost_formula_rows(ws, data_rows: List[int], col_130_percent: int, labor_cost_col: int, outsource_labor_cost_col: int, total_cost_col: int) -> None:
    """발생금액 = 130% + 공임대 + 외주공임"""
    for r in data_rows:
        addr_130 = ws.cell(row=r, column=col_130_percent).coordinate
        addr_labor = ws.cell(row=r, column=labor_cost_col).coordinate
        addr_outsource = ws.cell(row=r, column=outsource_labor_cost_col).coordinate
        set_cell_value_safe(ws, r, total_cost_col, f"={addr_130}+{addr_labor}+{addr_outsource}")


# =========================================================
# 7) 구상금액 수식(데이터 행만)
# =========================================================
def set_chargeback_formula_rows(ws, data_rows: List[int], occ_col: int, rate_col: int, chb_col: int) -> None:
    for r in data_rows:
        occ_addr = ws.cell(row=r, column=occ_col).coordinate
        rate_addr = ws.cell(row=r, column=rate_col).coordinate
        set_cell_value_safe(ws, r, chb_col, f"={occ_addr}*({rate_addr}/100)")


# =========================================================
# 7) 아래 합계 행(SUM, 필터 무시)
# =========================================================
def add_sum_rows(ws, first_row: int, last_row: int, occ_col: int, chb_col: int) -> None:
    """
    하단 합계 행 추가
    Args:
        first_row: 데이터 시작 행
        last_row: 데이터 마지막 행 (병합 포함)
        occ_col: 발생금액 컬럼 (T열)
        chb_col: 구상금액 컬럼 (V열)
    """
    # 1. 마지막 행 + 1행에 T열, V열 SUM 추가
    sum_row = last_row + 1
    
    # T열 SUM (발생금액) - 값이 표시되도록 명시적으로 설정
    sum_range_occ = f"{ws.cell(row=first_row, column=occ_col).coordinate}:{ws.cell(row=last_row, column=occ_col).coordinate}"
    sum_cell_occ = ws.cell(row=sum_row, column=occ_col)  # 병합 처리 없이 직접 접근
    sum_cell_occ.value = f"=SUM({sum_range_occ})"
    source_cell_occ = ws.cell(row=first_row, column=occ_col)
    sum_cell_occ.number_format = source_cell_occ.number_format if source_cell_occ.number_format else "_ * #,##0.00_ ;_ * -#,##0.00_ ;_ * \"-\"??_ ;_ @_"
    
    # V열 SUM (구상금액) - 값이 표시되도록 명시적으로 설정
    sum_range_chb = f"{ws.cell(row=first_row, column=chb_col).coordinate}:{ws.cell(row=last_row, column=chb_col).coordinate}"
    sum_cell_chb = ws.cell(row=sum_row, column=chb_col)  # 병합 처리 없이 직접 접근
    sum_cell_chb.value = f"=SUM({sum_range_chb})"
    source_cell_chb = ws.cell(row=first_row, column=chb_col)
    sum_cell_chb.number_format = source_cell_chb.number_format if source_cell_chb.number_format else "_ * #,##0.00_ ;_ * -#,##0.00_ ;_ * \"-\"??_ ;_ @_"
    
    # 2. 2행 띄우고 (마지막 행 + 3행) 발생금액/구상금액 레이블 추가
    label_start_row = last_row + 3
    
    # 발생금액
    set_cell_value_safe(ws, label_start_row, occ_col - 1, "발생금액")  # S열
    label_cell_occ = _cell_safe(ws, label_start_row, occ_col)  # T열
    label_cell_occ.value = f"={sum_cell_occ.coordinate}"  # 위에서 계산한 SUM 참조
    label_cell_occ.number_format = source_cell_occ.number_format if source_cell_occ.number_format else "_ * #,##0.00_ ;_ * -#,##0.00_ ;_ * \"-\"??_ ;_ @_"
    
    # 구상금액 (빨간색)
    label_cell_chb_text = _cell_safe(ws, label_start_row + 1, occ_col - 1)  # S열
    label_cell_chb_text.value = "구상금액"
    label_cell_chb_text.font = Font(color="FF0000")  # 빨간색 텍스트
    
    label_cell_chb = _cell_safe(ws, label_start_row + 1, occ_col)  # T열
    label_cell_chb.value = f"={sum_cell_chb.coordinate}"  # 위에서 계산한 SUM 참조
    label_cell_chb.number_format = source_cell_chb.number_format if source_cell_chb.number_format else "_ * #,##0.00_ ;_ * -#,##0.00_ ;_ * \"-\"??_ ;_ @_"
    label_cell_chb.font = Font(color="FF0000")  # 빨간색 숫자


# =========================================================
# 8) 상단 서브토탈(SUBTOTAL 9, 필터 반영)
# =========================================================
def set_subtotal_if_empty(ws, target_col: int, first_row: int, last_row: int, subtotal_row: int) -> None:
    """
    상단 서브토탈 설정
    Args:
        first_row: 데이터 시작 행
        last_row: 데이터 마지막 행 (병합 포함)
    """
    cell = ws.cell(row=subtotal_row, column=target_col)
    if not _is_blank(cell.value):
        return

    subtotal_range = f"{ws.cell(row=first_row, column=target_col).coordinate}:{ws.cell(row=last_row, column=target_col).coordinate}"
    subtotal_cell = _cell_safe(ws, subtotal_row, target_col)
    subtotal_cell.value = f"=SUBTOTAL(9,{subtotal_range})"
    # 원본 셀의 형식 참고 (회계 형식)
    source_cell = ws.cell(row=first_row, column=target_col)
    subtotal_cell.number_format = source_cell.number_format if source_cell.number_format else "_ * #,##0.00_ ;_ * -#,##0.00_ ;_ * \"-\"??_ ;_ @_"



# =========================================================
# 9) 룰 매칭 로직
# =========================================================
def check_rule_match(rule: Dict[str, Any], row_data: Dict[str, Any], current_date: str) -> bool:
    """
    룰이 현재 행에 적용 가능한지 확인
    
    Args:
        rule: 룰 데이터 (DB에서 조회한 딕셔너리)
        row_data: 현재 행 데이터 (vehicle, project_code, part_no, part_name, engine_form)
        current_date: 현재 날짜 (YYYY-MM-DD)
    
    Returns:
        True if 룰 적용 가능, False otherwise
    """
    # 1. 유효 기간 체크
    valid_from = rule.get("valid_from")
    valid_to = rule.get("valid_to")
    
    if valid_from and current_date < valid_from:
        return False
    if valid_to and current_date > valid_to:
        return False
    
    # 2. 프로젝트 코드 체크
    rule_project_code = rule.get("project_code", "ALL")
    if rule_project_code != "ALL":
        if row_data.get("project_code") != rule_project_code:
            return False
    
    # 3. 제외 프로젝트 코드 체크
    exclude_project_code = rule.get("exclude_project_code")
    if exclude_project_code:
        if row_data.get("project_code") == exclude_project_code:
            return False
    
    # 4. 차계 체크
    rule_vehicle = rule.get("vehicle_classification", "ALL")
    if rule_vehicle != "ALL":
        if row_data.get("vehicle") != rule_vehicle:
            return False
    
    # 5. 부품 체크 (부품번호 우선, 부품명 차선)
    rule_part_no = rule.get("part_no", "ALL")
    rule_part_name = rule.get("part_name", "ALL")
    
    # 룰에 부품번호나 부품명이 설정되어 있는 경우
    if rule_part_no != "ALL" or rule_part_name != "ALL":
        row_part_no = str(row_data.get("part_no", "")).strip()
        row_part_name = str(row_data.get("part_name", "")).strip()
        
        # 둘 다 설정된 경우: 부품번호 우선
        if rule_part_no != "ALL" and rule_part_name != "ALL":
            # 부품번호 매칭 시도
            if row_part_no and rule_part_no in row_part_no:
                pass  # 매칭 성공
            # 부품번호가 없거나 매칭 실패 시 부품명으로 시도 (대소문자 무시)
            elif row_part_name and rule_part_name.lower() in row_part_name.lower():
                pass  # 매칭 성공
            else:
                return False
        
        # 부품번호만 설정된 경우
        elif rule_part_no != "ALL":
            if not row_part_no or rule_part_no not in row_part_no:
                return False
        
        # 부품명만 설정된 경우 (대소문자 무시)
        elif rule_part_name != "ALL":
            if not row_part_name or rule_part_name.lower() not in row_part_name.lower():
                return False
    
    # 6. 엔진 형태 체크
    rule_engine_form = rule.get("engine_form", "ALL")
    if rule_engine_form != "ALL":
        if row_data.get("engine_form") != rule_engine_form:
            return False
    
    return True


# =========================================================
# 10) 마일리지/보증기간 필터(데이터 행만)
# =========================================================
def apply_warranty_filters_ws(
    ws,
    header_row: int,
    data_rows: List[int],
    mileage_threshold: int,
    warranty_years: int,
    rate_col: int,
) -> set[int]:
    mileage_col = pick_mileage_col(ws, header_row)
    sale_col = find_col_by_keywords_ws(ws, header_row, ["sale date", "판매일", "sale"], mode="any")
    # 수리일자 - "repair" 제외
    repair_col = find_col_by_keywords_ws(ws, header_row, ["repair date", "수리일자"], mode="any")

    changed_rows: set[int] = set()

    for r in data_rows:
        mv = parse_int_like(ws.cell(row=r, column=mileage_col).value)
        if mv is not None and mv >= mileage_threshold:
            set_cell_fill_safe(ws, r, mileage_col, FILL_HIGHLIGHT)
            set_rate(ws, r, rate_col, 0, changed_rows)

        sale_dt = parse_excel_date(ws.cell(row=r, column=sale_col).value)
        repair_dt = parse_excel_date(ws.cell(row=r, column=repair_col).value)
        if sale_dt and repair_dt:
            # 수리일에서 보증기간을 뺀 날짜 계산
            threshold_date = repair_dt - relativedelta(years=warranty_years)
            # 판매일이 threshold_date 이전이면 보증기간 초과
            if sale_dt < threshold_date:
                set_cell_fill_safe(ws, r, sale_col, FILL_HIGHLIGHT)
                set_rate(ws, r, rate_col, 0, changed_rows)

    for r in changed_rows:
        set_cell_fill_safe(ws, r, rate_col, FILL_HIGHLIGHT)

    return changed_rows


# =========================================================
# 11) 메인 처리(워크북 in-place) - 기존 레거시 함수
# =========================================================
@dataclass
class CompanyConfig:
    sheet_index: int = 0
    header_row: int = 3
    data_start_row: int = 4
    mileage_threshold: int = 50000
    warranty_years: int = 2
    anchor_keywords: Tuple[str, ...] = ("repair date", "수리일자", "repair")


def process_wb_inplace(wb: Workbook, cfg: CompanyConfig) -> None:
    """레거시 함수 - 룰 적용 없는 기본 전처리"""
    ws = wb.worksheets[cfg.sheet_index]

    # 차계 - vehicle 제외
    vehicle_col = find_col_by_keywords_ws(ws, cfg.header_row, ["vehicle classification", "차계"], mode="any")
    # 발생금액 - "발생" 제외
    occ_col = find_col_by_keywords_ws(ws, cfg.header_row, ["total cost", "발생금액"], mode="any")
    rate_col = find_rate_col(ws, cfg.header_row)
    chb_col = find_chargeback_col(ws, cfg.header_row)

    # 마지막 행 찾기: 교환부품번호 컬럼 사용 (병합 없음, 정확함)
    part_col = find_col_by_keywords_ws(ws, cfg.header_row, ["replaced part", "교환부품", "part no", "교환부품번호"], mode="any")
    last_row_guess = guess_last_data_row(ws, cfg.data_start_row, anchor_col=part_col, empty_run=30)

    unmerge_and_fill_column(ws, vehicle_col, cfg.data_start_row, last_row_guess)

    # 데이터 행 찾기는 기존 anchor_col 사용 (수리일자 등)
    anchor_col = find_col_by_keywords_ws(ws, cfg.header_row, list(cfg.anchor_keywords), mode="any")
    data_rows = iter_data_rows(ws, cfg.data_start_row, last_row_guess, anchor_col=anchor_col)
    if not data_rows:
        return

    apply_warranty_filters_ws(
        ws=ws,
        header_row=cfg.header_row,
        data_rows=data_rows,
        mileage_threshold=cfg.mileage_threshold,
        warranty_years=cfg.warranty_years,
        rate_col=rate_col,
    )

    set_chargeback_formula_rows(ws, data_rows, occ_col, rate_col, chb_col)
    # 수식 범위는 실제 마지막 행 사용 (병합 포함)
    add_sum_rows(ws, cfg.data_start_row, last_row_guess, occ_col, chb_col)

    # 상단 서브토탈: "구상금액" 기준
    subtotal_row = cfg.header_row - 1
    set_subtotal_if_empty(ws, target_col=chb_col, first_row=cfg.data_start_row, last_row=last_row_guess, subtotal_row=subtotal_row)
    
    # 자동 필터 설정 (3행 기준)
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{cfg.header_row}:{last_col_letter}{last_row_guess}"


# =========================================================
# 12) 새로운 전처리 함수 (룰 기반)
# =========================================================
def preprocess_with_rules(
    wb: Workbook,
    rule_table_name: str,
    repair_region: str,
    company_code: str = "",
    company_name: str = "",
    sheet_index: int = 0,
    header_row: int = 3,
    data_start_row: int = 4,
) -> PreprocessResult:
    """
    룰 기반 전처리 (새로운 구현)
    
    Args:
        wb: 워크북
        rule_table_name: 룰 테이블 이름 (예: "rule_company1")
        repair_region: 수리 지역 ("DOMESTIC" 또는 "OVERSEAS")
        company_code: 회사 코드
        company_name: 회사명
        sheet_index: 시트 인덱스
        header_row: 헤더 행 번호
        data_start_row: 데이터 시작 행 번호
    
    Returns:
        PreprocessResult: 전처리 결과 통계
    """
    result = PreprocessResult()
    result.repair_region = repair_region
    result.company_code = company_code
    result.company_name = company_name
    result.rule_table_name = rule_table_name
    result.process_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ws = wb.worksheets[sheet_index]
    current_date = datetime.now().strftime("%Y-%m-%d")
    
    # ===== 1단계: 컬럼 찾기 =====
    try:
        # 차계 - vehicle 제외
        vehicle_col = find_col_by_keywords_ws(ws, header_row, ["vehicle classification", "차계"], mode="any")
        
        # 부품번호 - 해외/국내 분리
        if repair_region == "OVERSEAS":
            part_no_col = find_col_by_keywords_ws(ws, header_row, ["PFP", "주원인부품번호"], mode="any")
        else:  # DOMESTIC
            part_no_col = find_col_by_keywords_ws(ws, header_row, ["replaced part", "교환부품", "part no", "교환부품번호"], mode="any")
        
        # 부품명
        part_name_col = find_col_by_keywords_ws(ws, header_row, ["part name", "부품명"], mode="any")
        
        # 엔진 형식 - engine form 추가
        engine_form_col = find_col_by_keywords_ws(ws, header_row, ["engine form", "엔진"], mode="any")
        
        rate_col = find_rate_col(ws, header_row)
        # 발생금액 - "발생" 제외
        occ_col = find_col_by_keywords_ws(ws, header_row, ["total cost", "발생금액"], mode="any")
        chb_col = find_chargeback_col(ws, header_row)
        
        # 공임비 및 발생금액 관련 컬럼 찾기
        col_130_percent = find_130_percent_col(ws, header_row)
        labor_cost_col = find_labor_cost_col(ws, header_row)
        outsource_labor_cost_col = find_outsource_labor_cost_col(ws, header_row)
        total_cost_col = occ_col  # occ_col과 동일한 컬럼
    except Exception as e:
        raise AppError(f"필수 컬럼을 찾을 수 없습니다: {e}")
    
    # ===== 2단계: 마지막 행 찾기 =====
    last_row_guess = guess_last_data_row(ws, data_start_row, anchor_col=part_no_col, empty_run=30)
    
    # ===== 3단계: 병합셀 해제 및 채우기 (차계) =====
    unmerge_and_fill_column(ws, vehicle_col, data_start_row, last_row_guess)
    
    # ===== 4단계: 데이터 행 찾기 =====
    # 수리일자 - "repair" 제외
    anchor_col = find_col_by_keywords_ws(ws, header_row, ["repair date", "수리일자"], mode="any")
    data_rows = iter_data_rows(ws, data_start_row, last_row_guess, anchor_col=anchor_col)
    
    if not data_rows:
        raise AppError("데이터 행을 찾을 수 없습니다.")
    
    result.total_rows = len(data_rows)
    
    # ===== 5단계: 전역 warranty 값 가져오기 =====
    global_mileage, global_warranty_years = _get_global_warranty()
    result.default_mileage_threshold = global_mileage
    result.default_warranty_years = global_warranty_years
    
    # ===== 6단계: 룰 필터링 (repair_region, status='ACTIVE', 날짜) =====
    all_rules = get_rules_from_table(rule_table_name)
    active_rules = [
        r for r in all_rules
        if r.get("status") == "ACTIVE"
        and r.get("repair_region") in ("ALL", repair_region)
    ]
    
    # 룰 사용 추적용 딕셔너리 초기화
    rule_used = {rule.get("rule_id"): False for rule in active_rules}
    
    # ===== 7단계: 발생금액 및 구상금액 수식 적용 =====
    # 발생금액 수식: 130% + 공임대 + 외주공임
    if total_cost_col and col_130_percent and labor_cost_col and outsource_labor_cost_col:
        set_total_cost_formula_rows(ws, data_rows, col_130_percent, labor_cost_col, outsource_labor_cost_col, total_cost_col)
    
    # 구상금액 수식: 발생금액 × 구상율 (나중에 후처리에서도 적용되지만, 여기서도 미리 적용)
    if chb_col and total_cost_col and rate_col:
        set_chargeback_formula_rows(ws, data_rows, total_cost_col, rate_col, chb_col)
    
    # ===== 8단계: 각 행 처리 =====
    # 행별 warranty 오버라이드 저장 {row_num: (mileage, years)}
    row_warranty_overrides: Dict[int, Tuple[Optional[int], Optional[int]]] = {}
    # 행별 룰 적용 여부 추적
    row_rule_applied: Dict[int, bool] = {}
    
    for row_num in data_rows:
        try:
            # 7-1. 차계 → 프로젝트 코드 추출
            vehicle_value = ws.cell(row=row_num, column=vehicle_col).value
            part_no_value = ws.cell(row=row_num, column=part_no_col).value
            project_code = get_project_code_from_vehicle(vehicle_value)
            
            vehicle_str = str(vehicle_value) if vehicle_value else ""
            part_no_str = str(part_no_value) if part_no_value else ""
            
            # 차계별 통계 추가
            if vehicle_str:
                result.add_vehicle_stat(vehicle_str, project_code)
            
            if project_code == "UNKNOWN":
                result.add_warning(row_num, vehicle_str, part_no_str, "프로젝트 코드 미매칭")
            
            # 7-2. common_project_liability 적용 (기본 구상률)
            liability_ratio = get_common_project_liability(project_code)
            base_liability_applied = False  # 기본 구상률 적용 여부
            if liability_ratio is not None:
                set_rate(ws, row_num, rate_col, liability_ratio, set())
                result.add_project_stat(project_code, liability_ratio)
                result.common_liability_applied += 1
                base_liability_applied = True  # 기본 구상률 적용됨
            else:
                result.add_project_stat(project_code, None)
                if project_code != "UNKNOWN":
                    result.add_warning(row_num, vehicle_str, part_no_str, "구상률 미설정")
            
            # 7-3. 행 데이터 구성
            # 부품명과 엔진 형식은 병합셀일 수 있으므로 좌상단 값 읽기
            top_row_part, top_col_part = _resolve_merged_anchor(ws, row_num, part_name_col)
            top_row_engine, top_col_engine = _resolve_merged_anchor(ws, row_num, engine_form_col)
            
            row_data = {
                "vehicle": vehicle_str,
                "project_code": project_code,
                "part_no": str(ws.cell(row=row_num, column=part_no_col).value or ""),
                "part_name": str(ws.cell(top_row_part, top_col_part).value or ""),
                "engine_form": str(ws.cell(top_row_engine, top_col_engine).value or ""),
            }
            
            # 필수 필드 누락 체크
            if not row_data["part_no"]:
                result.add_warning(row_num, vehicle_str, "", "부품번호 누락")
            
            # 7-4. 룰 매칭 및 적용 (우선순위 순)
            rule_applied = False  # 이 행에 룰이 적용되었는지 추적
            for rule in active_rules:
                if check_rule_match(rule, row_data, current_date):
                    rule_id = rule.get("rule_id")
                    rule_used[rule_id] = True
                    rule_applied = True  # 룰 적용됨
                    
                    # 룰 설명 생성 (format_rule_description 사용)
                    rule_desc = format_rule_description(rule)
                    result.add_rule_usage(rule_id, rule_desc)
                    
                    # Warranty 오버라이드 룰
                    warranty_mileage_override = rule.get("warranty_mileage_override")
                    warranty_period_override = rule.get("warranty_period_override")
                    
                    if warranty_mileage_override is not None or warranty_period_override is not None:
                        # 이미 저장된 오버라이드가 있어도 우선순위에 따라 덮어씀
                        current_override = row_warranty_overrides.get(row_num, (None, None))
                        new_mileage = warranty_mileage_override if warranty_mileage_override is not None else current_override[0]
                        new_years = warranty_period_override if warranty_period_override is not None else current_override[1]
                        row_warranty_overrides[row_num] = (new_mileage, new_years)
                        
                        if warranty_mileage_override is not None:
                            result.warranty_mileage_rules_applied += 1
                            result.add_mileage_override(warranty_mileage_override)
                        if warranty_period_override is not None:
                            result.warranty_period_rules_applied += 1
                            result.add_period_override(warranty_period_override)
                    
                    # 구상률 오버라이드 룰
                    rule_liability_ratio = rule.get("liability_ratio")
                    if rule_liability_ratio is not None:
                        set_rate(ws, row_num, rate_col, rule_liability_ratio, set())
                        result.liability_ratio_rules_applied += 1
                    
                    # 공임비 상한 룰
                    amount_cap_type = rule.get("amount_cap_type")
                    amount_cap_value = rule.get("amount_cap_value")
                    if amount_cap_type and amount_cap_type != "NONE" and amount_cap_value is not None:
                        # 공임비 컬럼 선택
                        target_col = None
                        if amount_cap_type == "LABOR" and labor_cost_col:
                            target_col = labor_cost_col
                        elif amount_cap_type == "OUTSOURCE_LABOR" and outsource_labor_cost_col:
                            target_col = outsource_labor_cost_col
                        elif amount_cap_type == "BOTH_LABOR" and total_cost_col:
                            target_col = total_cost_col
                        
                        if target_col:
                            try:
                                # 공임비 값 읽기
                                labor_value = ws.cell(row=row_num, column=target_col).value
                                if labor_value is not None:
                                    try:
                                        labor_amount = float(labor_value)
                                        # 상한 초과 체크
                                        if labor_amount > amount_cap_value:
                                            # 상한으로 제한
                                            set_cell_value_safe(ws, row_num, target_col, amount_cap_value)
                                            # 마킹 (노란색)
                                            set_cell_fill_safe(ws, row_num, target_col, FILL_HIGHLIGHT)
                                            result.amount_cap_exceeded_rows += 1
                                            result.amount_cap_rules_applied += 1
                                    except (ValueError, TypeError):
                                        pass  # 숫자 변환 실패 시 무시
                            except Exception:
                                pass  # 셀 읽기 실패 시 무시
            
            # 룰이 적용되었거나 기본 구상률이 적용된 경우 정상 처리로 카운트
            row_rule_applied[row_num] = rule_applied
            if rule_applied or base_liability_applied:
                result.success_rows += 1
            else:
                # 룰도 없고 기본 구상률도 없는 경우 예외 처리
                if not any(warning[0] == row_num for warning in result.warnings):
                    result.add_warning(row_num, vehicle_str, part_no_str, "적용 가능한 룰 및 기본 구상률 없음")
            
        except Exception as e:
            result.add_warning(row_num, vehicle_str if 'vehicle_str' in locals() else "?", part_no_str if 'part_no_str' in locals() else "?", f"처리 오류: {e}")
            result.error_rows += 1
    
    # 미사용 룰 기록
    for rule in active_rules:
        rule_id = rule.get("rule_id")
        if not rule_used.get(rule_id, False):
            # format_rule_description 사용
            rule_desc = format_rule_description(rule)
            
            reason = "조건 미매칭"
            if repair_region == "DOMESTIC" and rule.get("repair_region") == "OVERSEAS":
                reason = "지역 불일치 (해외 전용 룰)"
            elif repair_region == "OVERSEAS" and rule.get("repair_region") == "DOMESTIC":
                reason = "지역 불일치 (국내 전용 룰)"
            
            result.unused_rules.append((rule_id, rule_desc, reason))
    
    # ===== 8단계: Warranty 적용 =====
    mileage_col = pick_mileage_col(ws, header_row)
    sale_col = find_col_by_keywords_ws(ws, header_row, ["sale date", "판매일", "sale"], mode="any")
    # 수리일자 - "repair" 제외
    repair_col = find_col_by_keywords_ws(ws, header_row, ["repair date", "수리일자"], mode="any")
    
    for row_num in data_rows:
        # 오버라이드가 있으면 사용, 없으면 전역 값 사용
        if row_num in row_warranty_overrides:
            mileage_override, years_override = row_warranty_overrides[row_num]
            mileage_threshold = mileage_override if mileage_override is not None else global_mileage
            warranty_years = years_override if years_override is not None else global_warranty_years
        else:
            mileage_threshold = global_mileage
            warranty_years = global_warranty_years
        
        mileage_exceeded = False
        period_exceeded = False
        
        # 주행거리 체크
        mv = parse_int_like(ws.cell(row=row_num, column=mileage_col).value)
        if mv is not None and mv >= mileage_threshold:
            set_cell_fill_safe(ws, row_num, mileage_col, FILL_HIGHLIGHT)
            set_rate(ws, row_num, rate_col, 0, set())
            mileage_exceeded = True
            result.mileage_exceeded_rows += 1
        
        # 보증기간 체크
        sale_dt = parse_excel_date(ws.cell(row=row_num, column=sale_col).value)
        repair_dt = parse_excel_date(ws.cell(row=row_num, column=repair_col).value)
        if sale_dt and repair_dt:
            # 수리일에서 보증기간을 뺀 날짜 계산
            threshold_date = repair_dt - relativedelta(years=warranty_years)
            # 판매일이 threshold_date 이전이면 보증기간 초과
            if sale_dt < threshold_date:
                set_cell_fill_safe(ws, row_num, sale_col, FILL_HIGHLIGHT)
                set_rate(ws, row_num, rate_col, 0, set())
                period_exceeded = True
                result.period_exceeded_rows += 1
        
        if mileage_exceeded and period_exceeded:
            result.both_exceeded_rows += 1
        
        if mileage_exceeded or period_exceeded:
            set_cell_fill_safe(ws, row_num, rate_col, FILL_HIGHLIGHT)
            result.warranty_highlighted_rows += 1
    
    # ===== 9단계: 후처리 (구상금액, 합계) =====
    set_chargeback_formula_rows(ws, data_rows, occ_col, rate_col, chb_col)
    add_sum_rows(ws, data_start_row, last_row_guess, occ_col, chb_col)
    
    # 상단 서브토탈
    subtotal_row = header_row - 1
    set_subtotal_if_empty(ws, target_col=chb_col, first_row=data_start_row, last_row=last_row_guess, subtotal_row=subtotal_row)
    
    # 자동 필터 설정
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row_guess}"
    
    return result


# =========================================================
# 13) 파일 기반 처리(원하면 사용)
# =========================================================
def process_file(in_path: str, out_path: str, cfg: CompanyConfig) -> None:
    wb = load_workbook(in_path)
    process_wb_inplace(wb, cfg)
    wb.save(out_path)


def _infer_region_from_title(title: str) -> Optional[str]:
    """
    시트명으로 DOMESTIC/OVERSEAS 판단
    - '국내' 포함 => DOMESTIC
    - '해외' 포함 => OVERSEAS
    """
    if not title:
        return None
    t = str(title).strip().lower()
    if "국내" in t:
        return "DOMESTIC"
    if "해외" in t:
        return "OVERSEAS"
    return None


def _infer_region_with_opposite(wb: Workbook, idx: int, default_region: str = "") -> Optional[str]:
    """
    1) 현재 시트명으로 우선 판별
    2) 없으면 다른 시트들 중 '국내/해외'가 명확한 걸 찾아서 반대로 추정
       - 다른 시트가 DOMESTIC면 현재는 OVERSEAS로
       - 다른 시트가 OVERSEAS면 현재는 DOMESTIC로
    3) 그래도 없으면 default_region(기존 UI에서 넘어온 값) 사용
    """
    ws = wb.worksheets[idx]
    direct = _infer_region_from_title(ws.title)
    if direct:
        return direct

    other_regions = []
    for j, other_ws in enumerate(wb.worksheets):
        if j == idx:
            continue
        if other_ws.title == META_SHEET_NAME:
            continue
        r = _infer_region_from_title(other_ws.title)
        if r:
            other_regions.append(r)

    # 다른 시트에서 하나라도 잡히면 반대로 추정
    if "DOMESTIC" in other_regions and "OVERSEAS" not in other_regions:
        return "OVERSEAS"
    if "OVERSEAS" in other_regions and "DOMESTIC" not in other_regions:
        return "DOMESTIC"

    # 둘 다 섞여있거나 아예 없으면 default_region 사용(없으면 None)
    default_region = (default_region or "").strip().upper()
    if default_region in ("DOMESTIC", "OVERSEAS"):
        return default_region

    return None

# =========================================================
# 14) UI 엔트리 - 룰 기반 전처리
# =========================================================
def preprocess_inplace(
    wb: Workbook,
    company_code: str,
    company_name: str,
    rule_table_name: str,
    repair_region: str,  # ✅ UI에서 넘어오지만, 시트별 판별로 override 할거라서 "기본값" 취급
) -> PreprocessResult:
    """
    GUI 전처리 버튼 엔트리 (룰 기반) - 워크북 전체 시트 처리
    - _PREPROCESS_META는 스킵
    - 시트명에 '국내'/'해외' 포함이면 그걸로 repair_region 결정
    - 둘 다 없으면 반대 시트명으로 추정
    - 전처리 마킹은 마지막에 1번만
    """
    try:
        if _is_already_preprocessed(wb):
            raise AppError("이미 전처리된 파일입니다. (전처리는 1회만 가능합니다)")

        total = PreprocessResult()
        total.repair_region = repair_region
        total.company_code = company_code
        total.company_name = company_name
        total.rule_table_name = rule_table_name
        total.process_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for idx, ws in enumerate(wb.worksheets):
            if ws.title == META_SHEET_NAME:
                continue

            region = _infer_region_with_opposite(wb, idx, default_region=repair_region)
            if not region:
                total.add_remark(f"[{ws.title}] 스킵: 시트명에 국내/해외 없음 + 반대 추정 불가")
                continue

            r = preprocess_with_rules(
                wb=wb,
                rule_table_name=rule_table_name,
                repair_region=region,
                company_code=company_code,
                company_name=company_name,
                sheet_index=idx,
                header_row=3,
                data_start_row=4,
            )

            # ✅ 합산(최소 필드만)
            total.total_rows += r.total_rows
            total.success_rows += r.success_rows
            total.warning_rows += r.warning_rows
            total.error_rows += r.error_rows
            total.common_liability_applied += r.common_liability_applied
            total.warranty_mileage_rules_applied += r.warranty_mileage_rules_applied
            total.warranty_period_rules_applied += r.warranty_period_rules_applied
            total.liability_ratio_rules_applied += r.liability_ratio_rules_applied
            total.amount_cap_rules_applied += r.amount_cap_rules_applied
            total.amount_cap_exceeded_rows += r.amount_cap_exceeded_rows
            total.mileage_exceeded_rows += r.mileage_exceeded_rows
            total.period_exceeded_rows += r.period_exceeded_rows
            total.both_exceeded_rows += r.both_exceeded_rows
            total.warranty_highlighted_rows += r.warranty_highlighted_rows

            total.warnings.extend(r.warnings)
            total.remarks.extend([f"[{ws.title}] {m}" for m in r.remarks])

            if total.default_mileage_threshold == 0:
                total.default_mileage_threshold = r.default_mileage_threshold
            if total.default_warranty_years == 0:
                total.default_warranty_years = r.default_warranty_years

        _mark_preprocessed(wb)
        return total

    except AppError:
        raise
    except Exception as e:
        raise AppError(f"전처리 처리 중 오류: {e}") from e
